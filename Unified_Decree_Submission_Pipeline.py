#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Unified_Decree_Submission_Pipeline.py
==========================================================================
ONE script that replaces both:
    - Submit_Decree_Requests_Unified.py   (MDT creation + report + sign + merge)
    - Upload_Decree_PDFs.py               (final website upload of the merged PDF)

END-TO-END FLOW PER EXCEL ROW
--------------------------------------------------------------------------
  1. Create the MDT ("PreRequest") record on the SMC portal via the API
     calls the website itself makes -> get back the MDT request number
     (pre_request_id).
  2. Download the printable MDT form and stamp/sign it.
  3. Build the medical report from the Word template (wording depends on
     tumor type - see TUMOR_TYPE_CONFIG below).
  4. The patient document PDF is NOT looked up per-row anymore. Instead,
     BEFORE step 1 runs for ANY row, a batch pre-scan sweeps every unique
     patient ID in the whole queue at once (see "BATCH PRE-SCAN" below).
     By the time a given row reaches this step, its document has already
     been resolved (or the row has already been fully skipped, with no
     MDT even created — see step 4a).

--------------------------------------------------------------------------
BATCH PRE-SCAN (runs once, before any MDT creation for any row)
--------------------------------------------------------------------------
  a) For every unique patient ID in the queue, search
     D:\\SMS\\Cleaned_Processed_Patients_Doccuments (PATIENT_DOCS_ROOT,
     excluding its UNDER_PROCESSED subfolder) for the PDF named after
     that ID.
       - FOUND LOCALLY: check the CMIS archive
         (patient_pdf_dms_archive_fallback.py) for anything scanned in
         the last RECENT_ARCHIVE_DAYS_BACK days (default 7 — a full week,
         widened from the original today/yesterday-only window) that
         isn't already in this file — if found, merge it on and
         OVERWRITE the local file in place, then use it.
       - NOT found locally, the order is:
           i)  Try the SMC website first (patient_pdf_website_fallback.py).
               If found, ALSO merge in the last RECENT_ARCHIVE_DAYS_BACK
               days of CMIS archive pages, same as the local-found case —
               just applied to the freshly-downloaded file.
           ii) Only if the SMC website has nothing at all, fall back to
               the patient's ENTIRE CMIS archive
               (patient_pdf_dms_archive_fallback.py, full-extraction
               mode) as the final fallback.
       Anything freshly pulled by (i) or (ii) is saved into
       PATIENT_DOCS_UNDER_PROCESSED_DIR (a subfolder of PATIENT_DOCS_ROOT)
       — NOT directly into the cleaned root — so it's easy to find and is
       never mistaken for an already-cleaned local file.
       Only if BOTH fallbacks come up empty does that patient end up with
       no document at all (any row for that ID is later logged FAILED
       for that reason — this never blocks the rest of the run).
  b) If step (a) put anything into PATIENT_DOCS_UNDER_PROCESSED_DIR, your
     labeling server (DOC_LABELER_BATCH_PATH / start3.bat) is launched
     automatically, and the whole run PAUSES for you to clean/label those
     file(s) in the browser. Pressing Enter re-scans PATIENT_DOCS_ROOT for
     the now-cleaned copies (start3.bat's own --output already points
     there).
  c) ONE global pause (not per-row anymore) — only shown if step (a)
     found anything to freshly extract:
         'c' = continue  -> process EVERY row this run, including the
                             ones whose document had to be freshly
                             extracted/labeled.
         's' = skip      -> rows whose document had to be freshly
                             extracted this run are skipped ENTIRELY (no
                             MDT is even created for them); only rows
                             whose document was already local from the
                             start are processed.
     If every patient's document was already local, this pause never
     appears — there's nothing to decide.
  5. Merge: signed MDT form + medical report + the patient's full
     document PDF, in that order, into one PDF.
  6. Upload that merged PDF to the website using the same request flow
     the browser uses, and capture the final success message containing
     the decree request number ("رقم الطلب"). If any step fails, the row
     is logged FAILED with the step + reason; otherwise it's logged
     SUCCESS with both request numbers (MDT number + final decree number).

--------------------------------------------------------------------------
INPUT EXCEL  (row 1 = first patient, no header row assumed)
--------------------------------------------------------------------------
  Column A : patient national ID (SSN)
  Column B : treatment-plan / decree description text
  Column C : tumor type (OPTIONAL). Leave blank for the ordinary case
             (Breast Cancer, ~90-95% of rows). If present, must match one
             of the entries in TUMOR_TYPE_ALIASES / TUMOR_TYPE_CONFIG
             below (case-insensitive; matches either the Arabic diagnosis
             name or its diag code, e.g. "C25.9" or "سرطان البنكرياس").
             ~40 tumor types are pre-registered - see TUMOR_TYPE_CONFIG.
             Anything else is logged FAILED asking you to add a mapping,
             rather than guessing at diagnosis codes.
  Column D : request category (OPTIONAL). Leave blank for an ordinary
             request (uses the tumor type's own proc_id, e.g. 806 for
             breast / 10 for everything else). "surgery" forces proc_id
             486; "scan" forces proc_id 141 - in both cases the tumor
             type's diag_code/speciality_code are UNCHANGED, only
             proc_id is affected. See REQUEST_CATEGORY_ALIASES below.

--------------------------------------------------------------------------
TUMOR TYPE -> WHAT CHANGES
--------------------------------------------------------------------------
For every non-breast-cancer tumor type, three things change from the
ordinary workflow:
  a) The medical-report opening statement (native Arabic phrase for
     "Blood Type Tumor" as you supplied it; the other three currently
     hold the English phrases you gave verbatim as placeholders - see the
     "NOTE" comments in TUMOR_TYPE_CONFIG, replace with the exact Arabic
     wording you want before relying on this in production).
  b) GetDiagnosisData's specialityCode: 11 (breast) vs 9 (everything else
     covered here).
  c) The diagnosis code (INITIALICD10CODE) used throughout the MDT
     creation calls: AB45.6 (breast), C95 (blood), C20 (colorectal),
     C16.9 (gastric), C62.9 (testicular).
  d) GetReqTreatmentProcDesc's procId: 806 (breast) vs 10 (everything
     else covered here).
All other steps are identical regardless of tumor type.

--------------------------------------------------------------------------
RESUMABILITY / STABILITY
--------------------------------------------------------------------------
* Pause-after-first-row: the script fully completes row 1, then pauses
  for you to check it manually. Press Enter to continue with the rest,
  or Ctrl+C to stop and go fix whatever's wrong.
* Brief internet interruptions (a request-level connection error/timeout)
  during a network step: the script re-logs in and retries ONLY the
  in-progress step.
    - If the interruption happens during MDT creation, that whole step is
      restarted from scratch on retry (nothing from the half-finished
      attempt is reused - the site never saw a completed submission).
    - If the interruption happens after the MDT form was already fully
      submitted (i.e. you already have a pre_request_id), the script does
      NOT create a second MDT record - it resumes at signing/merging/
      uploading using the pre_request_id it already has.
* Crash / restart resilience: progress is also checkpointed to a small
  JSON file next to the run log after each major milestone (MDT created,
  merged PDF built, upload succeeded). If you stop the whole script and
  run it again on the same Excel file, already-completed rows are
  skipped and partially-completed rows resume from their last checkpoint
  instead of starting over.
* A row's medical-report + signed-MDT + patient-doc PDF is written to
  disk BEFORE upload is attempted, so a failed upload never loses that
  work - re-running the row re-uses the already-merged PDF.

==========================================================================
HOW TO RUN
==========================================================================
1. Install requirements once (same as the two scripts this replaces):
       pip install requests beautifulsoup4 openpyxl PyPDF2 pillow reportlab
       pip install pymupdf pdfkit arabic-reshaper python-bidi
   Also install wkhtmltopdf (see WKHTMLTOPDF_PATH note below). Word/COM
   is NOT required - the medical report is built directly as a PDF by
   medical_report_overlay.py (see that file's header for one-time setup).

2. Keep these three files in the SAME FOLDER as this script - they are
   imported directly, not standalone tools anymore:
       medical_report_overlay.py             (Stage 3 - report PDF)
       patient_pdf_dms_archive_fallback.py   (Stage 4 - CMIS archive; runs first)
       patient_pdf_website_fallback.py       (Stage 4 - SMC website; final fallback)
   The CMIS archive module needs HMIS_USERNAME / HMIS_PASSWORD set in
   the environment (same variables Extract_DMS_Patients_Prescriptioned_
   Services_Data.py uses). It looks patients up by national ID directly
   - no separate MR mapping file needed.

   BEFORE relying on it in a real run, test it standalone first:
       python dms_archive_explorer.py
   against one patient you know has archived documents, and confirm the
   PDF URLs it finds actually open. See that script's own header.

3. Edit the CONFIGURATION section below.

4. BEFORE running against real patients/live sites, test the newly
   added logic (flipped fallback order + the "review before continuing"
   pause) completely offline:
       python Unified_Decree_Submission_Pipeline.py --self-test
   This runs run_self_test() (near the bottom of this file), which
   exercises locate_patient_document_pdf() and
   prompt_review_extracted_pdf() with fake stand-ins for the network
   calls and for input() — it never touches CMIS, the SMC website, or
   your real files. It checks, and prints PASS/FAIL for, every branch:
   local-found (no pause), website-found with a CMIS same-day merge,
   website-found with nothing new in CMIS, website-empty but CMIS
   full-archive succeeds, and nothing found anywhere — plus that typing
   'c' resumes and 's' skips. Do not trust the real run until every
   line says PASS.

5. Run for real:
       python Unified_Decree_Submission_Pipeline.py
==========================================================================
"""

import glob
import io
import json
import logging
import os
import re
import sys
import shutil
import subprocess
import tempfile
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests
from bs4 import BeautifulSoup
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from PIL import Image, ImageEnhance, ImageFilter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import fitz  # PyMuPDF
import pdfkit
# NOTE: no docx2pdf / win32com / Word dependency at all anymore. Medical
# report generation is handled by medical_report_overlay.py, imported
# further below, which draws text directly onto a pre-exported PDF.

# =====================================================================
# CONFIGURATION - edit before running
# =====================================================================

BASE_URL = "https://smc.smcegy.com"
# Read from env, never hardcoded — a real password used to sit here in
# plaintext. decree_submission_service.py / decree_status_and_letters_sync.py
# already override these two module attributes from SMC_USERNAME/SMC_PASSWORD
# (or SMC_USERNAME_2/SMC_PASSWORD_2) before calling login(), so this change
# doesn't affect how either script behaves — it only removes the live
# credential that would otherwise sit in this file the moment it's committed.
USERNAME = os.environ.get("SMC_USERNAME", "")
PASSWORD = os.environ.get("SMC_PASSWORD", "")

# Input queue: Col A = ID, Col B = treatment plan, Col C = tumor type (optional).
SUBMISSION_ROOT = r"D:\SMC\decree_requesting_18-08_till_25_08"
EXCEL_INPUT_PATH = os.path.join(SUBMISSION_ROOT, "TEMPLATE_EXCEL_FILES\IDS_1.xlsx")

# Folder holding the full, already-cleaned per-patient PDF documents,
# searched by patient ID appearing anywhere in the filename.
PATIENT_DOCS_ROOT = r"D:\SMS\Cleaned_Processed_Patients_Doccuments"

# Freshly-extracted-but-not-yet-cleaned patient PDFs (SMC website / CMIS
# archive fallback) now land HERE instead of directly in PATIENT_DOCS_ROOT,
# so they're easy to find and don't get treated as "already cleaned" by
# find_patient_id_pdf() before you've had a chance to run them through the
# labeling server. This is exactly the folder your labeler's start3.bat
# already points --input at.
PATIENT_DOCS_UNDER_PROCESSED_DIR = os.path.join(PATIENT_DOCS_ROOT, "UNDER_PROCESSED")

# How many days back (inclusive of today) the CMIS archive is checked for
# "recently added investigation papers" to merge onto a patient's document
# PDF - whether that PDF was already local or was just pulled from the SMC
# website. Widened from the original today/yesterday (2-day) window to a
# full week, per your latest instruction.
RECENT_ARCHIVE_DAYS_BACK = 7

# Your local document-labeling/cleaning server (the one wired up via
# app.py / index.html / start3.bat). start3.bat already points --input at
# PATIENT_DOCS_UNDER_PROCESSED_DIR and --output at PATIENT_DOCS_ROOT, so no
# other wiring is needed for it to pick up freshly-extracted PDFs and save
# cleaned ones back to the folder find_patient_id_pdf() searches.
DOC_LABELER_BATCH_PATH = r"D:\Patients_Archived_Doccuments_Labeling_Server\medical_pdf_labeler\start3.bat"

# If True, the pipeline launches DOC_LABELER_BATCH_PATH for you
# automatically once the pre-scan finds anything to label. If False (or
# the path doesn't exist), it just prints where the file is and waits for
# you to start the labeler yourself.
LAUNCH_DOC_LABELER_AUTOMATICALLY = True

# Medical report template: the ORIGINAL .docx is kept only as the source
# you edit when wording changes. At runtime the pipeline uses a PDF
# exported from it once (see medical_report_overlay.py header for the
# one-time setup) - that PDF path is MEDICAL_REPORT_TEMPLATE_PDF there.
MEDICAL_REPORT_TEMPLATE = Path(r"D:\MDT_Medical_Report_Template\medical_report_template.docx")

# wkhtmltopdf binary - install once, then point this at it.
WKHTMLTOPDF_PATH = r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"
# Maximum time allowed for one authenticated MDT render.
RENDER_TIMEOUT_SECONDS = 90

# Signature / stamp images.
SIGNATURE_FILES = {
    "sig1": r"C:\Users\drmah\Template_Signatures\mdt1.png",
    "sig2": r"C:\Users\drmah\Template_Signatures\mdt2.png",
    "sig3": r"C:\Users\drmah\Template_Signatures\mdt3.png",
    "sig4": r"C:\Users\drmah\Template_Signatures\mdt4.png",
    "stamp": r"C:\Users\drmah\Template_Signatures\stamp.png",
}
SIGNATURES = [
    {"name": "Signature 1 - د/محمد كيلاني", "file_key": "sig1", "x": 85, "y": 475, "width": 120, "height": 50},
    {"name": "Signature 2 - ندى حسن", "file_key": "sig2", "x": 84, "y": 455, "width": 120, "height": 50},
    {"name": "Signature 3 - نرمين رمضان", "file_key": "sig3", "x": 81, "y": 440, "width": 120, "height": 50},
    {"name": "Signature 4 - ساره (Social Worker)", "file_key": "sig4", "x": 380, "y": 575, "width": 120, "height": 50},
]
STAMP = {"name": "Official Stamp - GRI-E", "file_key": "stamp", "x": 85, "y": 380, "width": 150, "height": 150}

SPECIALCOMMITTEEDOCTOR1 = "د/محمد كيلاني"
SPECIALCOMMITTEEDOCTOR2 = "د/ندى حسن"
SPECIALCOMMITTEEDOCTOR3 = "د/نرمين رمضان"

REQUEST_IMPORTANCE_ID = "22"   # "طلب عادي" (ordinary request)
GENDER_MAP = {"M": "1", "F": "2"}

PAUSE_AFTER_FIRST_ROW = True

# Network-interruption handling for API/upload steps.
NETWORK_MAX_RETRIES = 5
NETWORK_RETRY_WAIT_SECONDS = 5

# Cairo is UTC+3; the site encodes committee-date /Date(ms)/ timestamps
# as Cairo local midnight, not UTC midnight (see parse_committee_date).
CAIRO_UTC_OFFSET_HOURS = 3

# =====================================================================
# TUMOR TYPE CONFIGURATION
# =====================================================================
# Keyed by canonical id. Each entry:
#   label              - short English label, used in logs/run-log only
#   arabic_name         - official Arabic diagnosis name. Used both for
#                          alias matching AND to build the generic report
#                          phrase (see "opening_statement" below).
#   speciality_code     - GetDiagnosisData's specialityCode
#   diag_code           - INITIALICD10CODE used throughout MDT creation
#   proc_id             - ordinary/default GetReqTreatmentProcDesc procId.
#                          Overridden per-row by REQUEST_CATEGORY (surgery
#                          -> 486, scan -> 141) - see resolve_request_category
#                          below. diag_code/speciality_code are NEVER
#                          overridden by request category, only proc_id.
#   opening_statement    - the literal Arabic clinical opening line drawn
#                          into the report (treatment plan is appended
#                          after it). None => auto-generate the uniform
#                          format you asked for:
#                              "A patient of <arabic_name> for <plan>"
#                          Only Breast Cancer (matches the template's own
#                          baked-in wording) and Blood Type Tumor (real
#                          clinical phrase you supplied) have an explicit
#                          override; every other type uses the generic
#                          format below.
#
# NOTE ON DEFAULTS: speciality_code "9" and proc_id "10" are reused from
# the original four hand-configured types for every newly added type
# below, since that's what "everything else" used previously. This is an
# ASSUMPTION, not confirmed per-organ-system against the SMC system - if
# any of these types actually need a different speciality/committee
# routing, update that entry's speciality_code before relying on it in
# production (same caution as the original placeholder phrases had).
TUMOR_TYPE_CONFIG = {
    "breast_cancer": {
        "label": "Breast Cancer",
        "arabic_name": "سرطان الثدي",
        "speciality_code": "11",
        "diag_code": "AB45.6",
        "proc_id": "806",
        "opening_statement": "المريضة تعاني من ورم غير حميد بالثدى و تحتاج الي",
    },
    "blood_tumor": {
        "label": "Blood Type Tumor",
        "arabic_name": "لوكيميا ليمفاويه حاده",
        "speciality_code": "9",
        "diag_code": "C95",
        "proc_id": "10",
        "opening_statement": "المريض يعاني من لوكيميا ليمفاويه حاده ويحتاج الي",
    },
}


def _add_generic_tumor_type(canonical_id: str, english_label: str, arabic_name: str,
                             diag_code: str, speciality_code: str = "9", proc_id: str = "10",
                             extra_aliases: Optional[List[str]] = None):
    """Registers a tumor type that uses the uniform generic report format
    ("A patient of <english_label> for <treatment plan>") rather than a
    hand-supplied clinical phrase. This is what almost all of the newly
    added types below use.

    english_label is hand-written (not auto-derived from canonical_id) so
    it reads naturally inside that sentence - e.g. "head and neck cancer",
    not "Head Neck Cancer". It is used lower-cased at render time, so
    write it in whatever case reads best standalone (Title Case is fine)."""
    TUMOR_TYPE_CONFIG[canonical_id] = {
        "label": english_label,
        "arabic_name": arabic_name,
        "speciality_code": speciality_code,
        "diag_code": diag_code,
        "proc_id": proc_id,
        "opening_statement": None,  # -> generic "A patient of X for" format
    }
    aliases = {arabic_name.strip(): canonical_id, diag_code.strip().lower(): canonical_id,
               english_label.strip().lower(): canonical_id}
    for a in (extra_aliases or []):
        aliases[a.strip().lower()] = canonical_id
    _EXTRA_ALIASES.update(aliases)


_EXTRA_ALIASES: Dict[str, str] = {}

# Existing three (previously hand-written placeholder phrases) - migrated
# to the same uniform generic format as everything else below, now using
# their real Arabic diagnosis names instead of English placeholder text.
_add_generic_tumor_type("colorectal_cancer", "Colorectal Cancer", "سرطان القولون والمستقيم", "C20")
_add_generic_tumor_type("gastric_cancer", "Gastric Cancer", "سرطان المعدة", "C16.9",
                         extra_aliases=["stomach cancer"])
_add_generic_tumor_type("testicular_cancer", "Testicular Cancer", "سرطان الخصية", "C62.9",
                         extra_aliases=["testis cancer"])

# Explicitly requested addition - same mechanism as blood/colorectal etc,
# just pointed at pancreatic cancer.
_add_generic_tumor_type("pancreatic_cancer", "Pancreatic Cancer", "سرطان البنكرياس", "C25.9",
                         extra_aliases=["pancreas cancer"])

# --- Batch 2 of explicitly requested additions -------------------------
# english_label is set to EXACTLY the spelling you said will be typed in
# Column C of the Excel file, so it auto-registers as the matching alias
# with no extra_aliases needed (resolve_tumor_type() only lower-cases and
# collapses whitespace - it does not fix spelling/spacing on its own).
#
# CAUTION - Arabic diagnosis names below are my best direct translation
# of each English label, NOT text you supplied or that's been confirmed
# against the SMC/CMIS system (same caution the file's original
# placeholder-phrase types called out). These are only used to build the
# generic report sentence ("A patient of <arabic_name> for <plan>") and
# for GetDiagnosisData alias matching - verify each one before relying on
# it in a real submission, and replace with the official wording if it
# differs.
#
# "abdominal_pelvic_cancer" reuses diag code U55.5 - the SAME code
# already used by the existing "pelvic_cancer" entry (سرطان بالحوض).
# Both are kept as separate, independently-selectable types (same
# pattern as the existing colon_cancer/colorectal_cancer split above) -
# Column C text decides which one a row resolves to.
#
# "spreaded_tumor" reuses diag code C80 - the SAME code that was
# EXPLICITLY DROPPED earlier in this file as "cancer_unknown_primary"
# (see the note above _SUPPLIED_TUMOR_TABLE). That old drop is now
# effectively superseded for C80: any row whose Column C says "Spreaded
# Tumor" WILL resolve and process (via this new entry) - C80 is no
# longer a guaranteed-fail code for every row, only for text that
# doesn't match "Spreaded Tumor" (or any other alias you add for it).
_add_generic_tumor_type("abdominal_cancer", "Abdominal Cancer", "سرطان بالبطن", "C49.0")
_add_generic_tumor_type("abdominal_pelvic_cancer", "Abdominal & Pelvic Cancer", "سرطان بالبطن والحوض", "U55.5")
_add_generic_tumor_type("neck_cancer", "Neck Cancer", "سرطان بالرقبه", "AB43.8")
_add_generic_tumor_type("spreaded_tumor", "Spreaded Tumor", "ثانويات سرطانيه غير معلومه المصدر", "C80")
_add_generic_tumor_type("arm_cancer", "Arm Cancer", "سرطان بالذراع", "AB38.6")

# --- Batch 3: previously-missing app-module types, codes you supplied --
_add_generic_tumor_type("prostate_cancer", "Prostate Cancer", "سرطان البروستاتا", "C61")
_add_generic_tumor_type("lymphoma", "Lymphoma", "سرطان الغدد الليمفاوية", "C77",
                         # "اورام ليمفاويه" is the exact cancer_type_group text the app
                         # itself uses for this diag code (per All_Diag_Codes.xlsx /
                         # the "امراض الدم" bucket) - added as an alias since alias
                         # matching is exact-string, not fuzzy, and differs from the
                         # arabic_name above.
                         extra_aliases=["اورام ليمفاويه"])
_add_generic_tumor_type("brain_cns", "Brain and Central Nervous System Cancer",
                         "أورام المخ والجهاز العصبي المركزي", "U37.6")
_add_generic_tumor_type("gynecological_cancer", "Gynecological Cancer", "أورام النساء", "C56")

# Explicitly requested addition - diag_code C56 as supplied. Reuses the
# SAME code as "gynecological_cancer" above (same intentional-code-reuse
# pattern already used elsewhere in this file for pelvic_cancer /
# abdominal_pelvic_cancer, both U55.5, and colon_cancer / colorectal_cancer
# both being distinct-but-related codes) - Column C text alone decides
# which of the two a row resolves to; they are independent, separately
# selectable entries. speciality_code/proc_id left at the "9"/"10"
# default per the note above _SUPPLIED_TUMOR_TABLE - update if this
# needs its own committee/proc routing.
_add_generic_tumor_type("ovarian_cancer", "Ovarian Cancer", "سرطان المبيض", "C56")

# Explicitly requested addition - Thyroid Gland Tumor, diag_code C73 as
# supplied. english_label is set to match the exact Column C spelling
# seen in the run log ("thyroid Gland Tumor") - resolve_tumor_type()
# lower-cases before matching, so the auto-derived alias
# ("thyroid gland tumor") covers that spelling regardless of case.
# A couple of likely spelling variants are added as extra_aliases up
# front (same defensive pattern used for the other types above) rather
# than waiting for another "Unrecognized tumor type" failure.
_add_generic_tumor_type("thyroid_cancer", "Thyroid Gland Tumor", "سرطان الغدة الدرقية", "C73",
                         extra_aliases=["thyroid cancer", "thyroid gland cancer",
                                        "thyroidcancer", "thyroid tumor"])

# --- The full list you supplied (code -> Arabic diagnosis name) --------
# canonical_id, english_label, arabic_name, diag_code. english_label is
# hand-written for natural phrasing inside "A patient of X for ..." -
# not a literal translation in every case, but the clinically-standard
# English name for that diagnosis. speciality_code/proc_id default to
# "9"/"10" per the note above unless a row overrides them.
#
# "cancer_unknown_primary" (C80, "ثانويات سرطانيه غير معلومه المصدر") from
# your original list has been dropped entirely per your instruction - it
# is not registered and rows using it will fail with a clear "unknown
# tumor type" message rather than silently matching something else.
_SUPPLIED_TUMOR_TABLE = [
    ("bone_cancer",                    "Bone Cancer",                              "سرطان بالعظام",                      "C40.9"),
    ("lip_cancer",                     "Lip Cancer",                               "سرطان الشفه",                        "C44.0"),
    ("soft_tissue_sarcoma",            "Soft Tissue Sarcoma",                      "سرطان بالانسجه الرخوه",              "C49.9"),
    ("pelvic_cancer",                  "Pelvic Cancer",                            "سرطان بالحوض",                       "U55.5"),
    ("clitoral_cancer",                "Clitoral Cancer",                          "سرطان بالبظر",                       "AA93.6"),
    ("pharyngeal_cancer",              "Pharyngeal Cancer",                        "سرطان بالبلعوم",                     "C13.9"),
    ("nasal_cavity_cancer",            "Nasal Cavity Cancer",                      "سرطان بداخل الانف",                  "C30.0"),
    ("sinus_cancer",                   "Paranasal Sinus Cancer",                   "سرطان بالجيوب الانفيه",              "C31.9"),
    ("laryngeal_cancer",               "Laryngeal Cancer",                         "سرطان بالحنجره",                     "C32.9"),
    ("head_neck_cancer",               "Head and Neck Cancer",                     "سرطان بالرأس والرقبه",               "C76.0"),
    ("breast_cancer_AB45.6",             "Breast Cancer",                            "سرطان الثدي",                        "AB45.6"),
    ("lung_cancer",                    "Lung Cancer",                              "سرطان الرئة",                        "C34.9"),
    ("bladder_cancer",                 "Bladder Cancer",                           "سرطان المثانه البوليه",              "C67.9"),
    ("kidney_cancer",                  "Kidney Cancer",                            "سرطان الكلي",                        "C64.9"),
    ("esophageal_cancer",              "Esophageal Cancer",                        "سرطان المرىء",                       "C15.9"),
    ("colon_cancer",                   "Colon Cancer",                             "سرطان القولون",                      "C18.9"),
    ("anal_cancer",                    "Anal Cancer",                              "سرطان بالشرج",                       "C21.0"),
    ("liver_cancer",                   "Liver Cancer",                             "سرطان الكبد",                        "C22.9"),
    ("bile_duct_cancer",               "Bile Duct Cancer",                         "سرطان القنوات المراريه",             "C24.9"),
    ("cervical_cancer",                "Cervical Cancer",                          "سرطان بعنق الرحم",                   "C53.9"),
    ("adrenal_gland_cancer",           "Adrenal Gland Cancer",                     "سرطان بالغده الكظريه",               "C74.9"),
    ("pleural_mesothelioma",           "Pleural Mesothelioma",                     "سرطان بالغشاء البللوري",             "C38.4"),
    ("skin_cancer",                    "Skin Cancer",                              "سرطان بالجلد",                       "C44.9"),
    ("tongue_cancer",                  "Tongue Cancer",                            "سرطان باللسان",                      "C02.9"),
    ("mouth_cancer",                   "Oral Cavity Cancer",                       "سرطان بالفم",                        "C04.9"),
    ("palate_cancer",                  "Palate Cancer",                            "سرطان سقف الفم",                     "C05.9"),
    ("salivary_gland_cancer",          "Salivary Gland Cancer",                    "سرطان بالغده اللعابيه",              "C08.9"),
    ("ureter_cancer",                  "Ureteral Cancer",                          "سرطان بالحالب",                      "AA51.9"),
    ("neuroblastoma",                  "Neuroblastoma",                            "سرطان بالعقد العصبيه",               "AB10.8"),
    ("hepatitis_c_liver_focus",        "Hepatitis C with Suspected Hepatic Focal Lesion", "التهاب كبدي سي واشتباه بؤره سرطانيه بالكبد", "AA59.2"),
    ("liver_cirrhosis_diabetes_focus", "Liver Cirrhosis with Suspected Hepatic Focal Lesion", "تليف كبدي وسكر بالدم وتضخم بالطحال واشتباه بؤره سرطانيه بالكبد", "U94.4"),
    ("plasma_cell_cancer",             "Plasma Cell Cancer",                       "سرطان بالخلايا البلازميه",           "AB23.9"),
    ("post_bladder_tumor_resection",   "Bladder Cancer, Status Post Resection",    "ما بعد استئصال ورم سرطاني بالمثانه", "AB12.8"),
    ("cancer_transfusion_need",        "Cancer Requiring Blood Transfusion",       "مريض سرطان يحتاج نقل دم ومشتقاته",   "AB18.9"),
    ("skin_cancer_mf",                 "Mycosis Fungoides",                        "سرطان بالجلد (M.F)",                 "U86.9"),
    ("neural_cell_cancer",             "Neural Cell Cancer",                       "سرطان بالخلايا العصبيه",             "AB27.6"),
    ("post_jaw_tumor_resection",       "Jaw Tumor, Status Post Resection with Bone Deformity", "ما بعد استئصال ورم سرطاني بالفك مع تشوه بعظام الفك", "AB30.0"),
    ("neuroendocrine_recurrent",       "Recurrent Treatment-Resistant Neuroendocrine Tumor", "ورم سرطاني بالغدد العصبيه مرتجع غير مستجيب للعلاج", "AB45.1"),
    ("rectal_cancer_stricture",        "Rectal Cancer with Colonic Stricture",     "سرطان مستقيم + ضيق بالقولون",        "AB31.5"),
]
for _cid, _elabel, _arname, _dcode in _SUPPLIED_TUMOR_TABLE:
    _add_generic_tumor_type(_cid, _elabel, _arname, _dcode)
del _cid, _elabel, _arname, _dcode

# NOTE on "colon_cancer" (C18.9) vs the pre-existing "colorectal_cancer"
# (C20): these are two distinct ICD codes from your table (colon only vs
# colon+rectum). Both are now registered as separate types. If Column C
# says "colon cancer" it now resolves to the more specific C18.9 entry;
# "colorectal cancer" still resolves to the original C20 entry. Verify
# this split matches what you actually want before relying on it.
#
# NOTE on "breast_cancer_c509": your table also lists سرطان الثدي / AB45.6
# as its own row, distinct from the primary "breast_cancer" entry (blank
# Column C default, diag AB45.6, which keeps the template's own baked-in
# wording). This AB45.6 entry is treated as a generic (non-default,
# masked + generic-phrase) type like everything else in the table, NOT
# as an alias for the default breast case - so it will NOT reuse the
# template's baked sentence, it gets the generic "A patient of سرطان
# الثدي for ..." phrasing instead. Adjust if you intended otherwise.

# --- Batch 4: added from All_Diag_Codes.xlsx (the app's full cancer-type /
# diag-code export) - these 17 diag codes appeared in that file but had no
# matching entry anywhere above yet. Same convention as every other batch:
# speciality_code/proc_id default to "9"/"10" (unconfirmed per-organ
# against SMC - override on a specific entry if it turns out to need its
# own committee routing), arabic_name is copied VERBATIM from
# All_Diag_Codes.xlsx (that's also the exact text the app's
# cancer_type_group column uses, so it doubles as the alias for
# JS-export Column C matching), english_label is hand-written since the
# source file only supplied Arabic names.
_ALL_DIAG_CODES_ADDITIONS = [
    ("breast_suspected_initiative",       "Suspected Breast Cancer Under Diagnosis (Initiative)",              "اشتباه سرطان ثدي تحت التشخيص - مبادرة", "AB42.5"),
    ("facial_tumor",                      "Facial Tumor",                                                      "اورام بالوجه",                          "C41.0"),
    ("jaw_tumor",                         "Jaw Tumor",                                                          "اورام بالفك",                           "C41.1"),
    ("chest_tumor",                       "Chest Tumor",                                                        "اورام بالصدر",                          "C39.9"),
    ("malignant_bone_marrow_tumor",       "Malignant Bone Marrow Tumor",                                        "اورام خبيثة بالنخاع",                   "C72"),
    ("anemia_suspected_tumor",            "Treatment-Resistant Iron-Deficiency Anemia with Suspected Tumor",   "انيميا نقص الحديد غير مستجيبه للعلاج واشتباة وجود اورام", "U82.9"),
    ("ear_tumor",                         "Ear Tumor",                                                          "اورام بالاذن",                          "D14.0"),
    ("leg_tumor",                         "Leg Tumor",                                                          "اورام بالساق",                          "C40.2"),
    ("eye_tumor",                         "Eye Tumor",                                                          "اورام بالعين",                          "C69.9"),
    ("gum_tumor",                         "Gum Tumor",                                                          "اورام باللثه",                          "C03.9"),
    ("malignant_endocrine_gland_tumor",   "Malignant Endocrine Gland Tumor",                                    "اورام خبيثة بالغدد الصماء",             "C75"),
    ("fibrous_sclerosis_brain_tumor",     "Fibrous Sclerosis with Brain Tumors",                                "تصلب تليفي مع اورام بالمخ",             "AB12.0"),
    ("liver_cirrhosis_ascites_tumor",     "Liver Cirrhosis with Ascites and Liver Tumors",                      "تليف كبدي+استسقاء+اورام كبد",           "AA58.9"),
    ("benign_bone_tumor",                 "Benign Bone Tumor",                                                  "اورام حميده بالعظام",                   "AA58.6"),
    ("abdominal_lymph_node_tumor",        "Abdominal Lymph Node Tumor",                                         "اورام بالعقد الليمفاويه بالبطن",        "AB29.9"),
    ("lung_cancer_early_detection",       "Early Lung Cancer Detection (Initiative)",                           "كشف مبكر اورام الرئه (مبادره)",         "AB51.7"),
    ("prostate_cancer_early_detection",   "Early Prostate Cancer Detection (Initiative)",                       "كشف مبكر اورام بروستاتا (مبادره)",      "AB51.6"),
]
for _cid, _elabel, _arname, _dcode in _ALL_DIAG_CODES_ADDITIONS:
    _add_generic_tumor_type(_cid, _elabel, _arname, _dcode)
del _cid, _elabel, _arname, _dcode

# NOTE on "colon_cancer" (C18.9) vs the pre-existing "colorectal_cancer"
# (C20): these are two distinct ICD codes from your table (colon only vs
# colon+rectum). Both are now registered as separate types. If Column C
# says "colon cancer" it now resolves to the more specific C18.9 entry;
# "colorectal cancer" still resolves to the original C20 entry. Verify
# this split matches what you actually want before relying on it.
#
# NOTE on "breast_cancer_c509": your table also lists سرطان الثدي / AB45.6
# as its own row, distinct from the primary "breast_cancer" entry (blank
# Column C default, diag AB45.6, which keeps the template's own baked-in
# wording). This AB45.6 entry is treated as a generic (non-default,
# masked + generic-phrase) type like everything else in the table, NOT
# as an alias for the default breast case - so it will NOT reuse the
# template's baked sentence, it gets the generic "A patient of سرطان
# الثدي for ..." phrasing instead. Adjust if you intended otherwise.

# --- App-module naming compatibility --------------------------------
# Your app's own TUMOR_TYPES module names the same diagnoses using
# different text (mostly SCREAMING_SNAKE_CASE constant names, e.g.
# "BONE_CANCER", "PARANASAL_SINUS_CANCER") than what's used above. Every
# one of these that came with its own ICD-10 code in your module already
# matches an existing entry above with THE SAME code, so no new diag
# codes were needed - these are alias-only additions so Column C accepts
# either naming and still resolves to the exact same canonical entry /
# diag_code / speciality_code / proc_id as before.
#
# Two of your module's UNCODED entries also turned out to already be
# covered above, just under more specific Arabic wording than your
# module uses - added as aliases to the SAME existing canonical entry
# and diag code (no code change): BLOOD_RELATED_TUMORS -> blood_tumor
# (C95) and BLADDER_CANCER -> bladder_cancer (C67.9).
#
# PROSTATE_CANCER (C61), LYMPHOMA (C77), BRAIN_CNS (U37.6), and
# GYNECOLOGICAL_CANCER (C56) were the previously-missing types - now
# registered above in "Batch 3" with the codes you supplied.
#
# "bone_soft_tissue" was a typo on the module side; the corrected module
# name "BONE_SOFT_TISSUE" is aliased below directly to the
# existing "soft_tissue_sarcoma" entry (C49.9) - no separate combined
# code was needed, per your correction.
_APP_MODULE_ALIASES = {
    # key (module constant name, both underscore and space forms are
    # registered) -> canonical id already in TUMOR_TYPE_CONFIG.
    "breast_cancer": "breast_cancer",
    "blood_related_tumors": "blood_tumor",
    "colorectal_cancer": "colorectal_cancer",
    "lung_cancer": "lung_cancer",
    "liver_cancer": "liver_cancer",
    "bladder_cancer": "bladder_cancer",
    "bone_cancer": "bone_cancer",
    "soft_tissue_sarcoma": "soft_tissue_sarcoma",
    "colon_cancer": "colon_cancer",
    "rectal_cancer_with_colonic_stricture": "rectal_cancer_stricture",
    "cervical_cancer": "cervical_cancer",
    "clitoral_cancer": "clitoral_cancer",
    "lip_cancer": "lip_cancer",
    "pelvic_cancer": "pelvic_cancer",
    "pharyngeal_cancer": "pharyngeal_cancer",
    "nasal_cavity_cancer": "nasal_cavity_cancer",
    "paranasal_sinus_cancer": "sinus_cancer",
    "laryngeal_cancer": "laryngeal_cancer",
    "head_and_neck_cancer": "head_neck_cancer",
    "kidney_cancer": "kidney_cancer",
    "esophageal_cancer": "esophageal_cancer",
    "anal_cancer": "anal_cancer",
    "bile_duct_cancer": "bile_duct_cancer",
    "adrenal_gland_cancer": "adrenal_gland_cancer",
    "pleural_mesothelioma": "pleural_mesothelioma",
    "skin_cancer": "skin_cancer",
    "tongue_cancer": "tongue_cancer",
    "oral_cavity_cancer": "mouth_cancer",
    "palate_cancer": "palate_cancer",
    "salivary_gland_cancer": "salivary_gland_cancer",
    "ureteral_cancer": "ureter_cancer",
    "neuroblastoma": "neuroblastoma",
    "hepatitis_c_with_suspected_hepatic_focal_lesion": "hepatitis_c_liver_focus",
    "liver_cirrhosis_with_suspected_hepatic_focal_lesion": "liver_cirrhosis_diabetes_focus",
    "plasma_cell_cancer": "plasma_cell_cancer",
    "bladder_cancer_status_post_resection": "post_bladder_tumor_resection",
    "cancer_requiring_blood_transfusion": "cancer_transfusion_need",
    "mycosis_fungoides": "skin_cancer_mf",
    "neural_cell_cancer": "neural_cell_cancer",
    "jaw_tumor_status_post_resection_with_bone_deformity": "post_jaw_tumor_resection",
    "recurrent_treatment_resistant_neuroendocrine_tumor": "neuroendocrine_recurrent",
    # Corrected module name (was the typo "bone_soft_tissue") -> points
    # to the existing Soft Tissue Sarcoma entry (C49.9), per your note.
    "BONE_SOFT_TISSUE": "soft_tissue_sarcoma",
    "prostate_cancer": "prostate_cancer",
    "lymphoma": "lymphoma",
    "brain_cns": "brain_cns",
    "gynecological_cancer": "gynecological_cancer",
}
for _mod_key, _canon in _APP_MODULE_ALIASES.items():
    # Register BOTH the raw underscore form ("bone_cancer") and a
    # space-separated form ("bone cancer") - resolve_tumor_type() only
    # lower-cases + collapses whitespace, it does NOT turn underscores
    # into spaces, so both forms need their own explicit entry here.
    _EXTRA_ALIASES.setdefault(_mod_key, _canon)
    _EXTRA_ALIASES.setdefault(_mod_key.replace("_", " "), _canon)
del _mod_key, _canon

# Free-text values from Excel Col C -> canonical TUMOR_TYPE_CONFIG key.
# Matching is case-insensitive with whitespace collapsed. Add more
# aliases here as new spellings/types show up in your data. Arabic
# diagnosis names and diag codes are auto-added for every type above via
# _add_generic_tumor_type(); this dict adds the hand-curated English
# synonyms for the original four types plus the blank-Column-C default.
TUMOR_TYPE_ALIASES = {
    "": "breast_cancer",
    "breast cancer": "breast_cancer",
    "breast": "breast_cancer",
    # Arabic name for plain breast cancer - hand-curated so it always
    # resolves to the primary breast_cancer entry (806/AB45.6/template
    # phrase), never to the separate breast_cancer_c509 entry below,
    # which happens to share the exact same Arabic diagnosis name.
    "سرطان الثدي": "breast_cancer",
    "blood type tumor": "blood_tumor",
    "blood tumor": "blood_tumor",
    "blood cancer": "blood_tumor",
    "leukemia": "blood_tumor",
    "leukaemia": "blood_tumor",
    # Hand-curated so these resolve correctly regardless of exact spacing
    # in Column C - resolve_tumor_type() only lower-cases + collapses
    # whitespace, it doesn't insert missing spaces, so "LungCancer" (no
    # space) would NOT otherwise match the auto-registered "lung cancer"
    # alias from lung_cancer's english_label. Same reasoning for
    # "livercancer". "soft tissue cancer" is a wording variant of
    # soft_tissue_sarcoma's own auto-registered alias ("soft tissue
    # sarcoma") and needs its own explicit entry for the same reason.
    "lungcancer": "lung_cancer",
    "livercancer": "liver_cancer",
    "soft tissue cancer": "soft_tissue_sarcoma",
    # Added after checking against the real IDS.xlsx: these no-space /
    # differently-punctuated spellings are exactly what Column C actually
    # contains for these rows, and none of them were otherwise reachable:
    # "coloncancer"/"cervicalcancer"/"prostatecancer"/"colorectalcancer"
    # simply had no no-space alias registered (same missing-alias reason
    # as lungcancer/livercancer above). "bonesofttissue" is a separate,
    # worse bug fix: the existing "BONE_SOFT_TISSUE" / "BONE SOFT TISSUE"
    # aliases (from _APP_MODULE_ALIASES) were registered in UPPERCASE,
    # but resolve_tumor_type() lower-cases its lookup key before checking
    # this dict - so those two aliases could never match ANY input,
    # correctly spaced or not. Adding the lowercase, no-space form here
    # (hand-curated dict always wins collisions - see the loop below)
    # fixes it for good instead of just patching this one spelling.
    "coloncancer": "colon_cancer",
    "cervicalcancer": "cervical_cancer",
    "prostatecancer": "prostate_cancer",
    "colorectalcancer": "colorectal_cancer",
    "bonesofttissue": "soft_tissue_sarcoma",
    "bone_soft_tissue": "soft_tissue_sarcoma",
    "bone soft tissue": "soft_tissue_sarcoma",
    # Added after checking every distinct Column C value actually present
    # in your real IDS.xlsx queue against resolve_tumor_type(): these were
    # silently falling through to "Unrecognized tumor type" because only
    # a spaced/underscored form was registered (via _APP_MODULE_ALIASES or
    # the auto-derived english_label alias), never the plain no-space
    # form your sheet actually types. Same missing-alias reason as
    # lungcancer/livercancer/coloncancer above - not a new tumor type,
    # just a missing spelling for one that already existed with a diag
    # code (canonical id / diag code noted per line for verification):
    "breastcancer": "breast_cancer",              # AB45.6
    "bonecancer": "bone_cancer",                  # C40.9
    "bloodrelatedtumors": "blood_tumor",           # C95
    "bladdercancer": "bladder_cancer",             # C67.9
    "kidneycancer": "kidney_cancer",               # C64.9
    "plasmacellcancer": "plasma_cell_cancer",      # AB23.9
    # "Brain Tumor" (the exact text used in your sheet) had NO alias at
    # all pointing at the existing brain_cns entry (U37.6) - its own
    # auto-derived alias only covers the full english_label "brain and
    # central nervous system cancer". Assumed to mean the same thing;
    # flag this specific mapping for a quick sanity check on your end.
    "brain tumor": "brain_cns",                    # U37.6
}

# FIXED BUG: this used to be a blind TUMOR_TYPE_ALIASES.update(_EXTRA_ALIASES),
# which silently let auto-derived aliases overwrite the hand-curated ones
# above. That's exactly why breast cancer was using proc_id 10 instead of
# 806: _add_generic_tumor_type("breast_cancer_c509", "Breast Cancer", ...)
# auto-registers "breast cancer" (its english_label, lower-cased) as an
# alias pointing at breast_cancer_c509 (diag AB45.6, proc_id 10, generic
# phrasing) - and since that update ran AFTER the hand-curated dict above,
# it silently clobbered the correct "breast cancer" -> "breast_cancer"
# entry for every row where Column C literally said "Breast Cancer" (the
# ~90-95% majority case). Blank Column C was unaffected (no auto-alias
# for ""), which is why this stayed hidden until you checked a row that
# had "Breast Cancer" typed explicitly.
# Now: hand-curated aliases always win. Any auto-derived alias that would
# collide with a DIFFERENT canonical id than an existing entry is skipped
# and printed here, instead of silently applied - so a future collision
# like this one is impossible to miss.
for _alias_key, _alias_target in _EXTRA_ALIASES.items():
    if _alias_key in TUMOR_TYPE_ALIASES and TUMOR_TYPE_ALIASES[_alias_key] != _alias_target:
        print(f"  [tumor alias collision - IGNORED] {_alias_key!r}: keeping "
              f"{TUMOR_TYPE_ALIASES[_alias_key]!r}, NOT overwriting with "
              f"auto-derived {_alias_target!r} (add an explicit override above "
              f"if {_alias_target!r} was actually what you wanted for this text)")
        continue
    TUMOR_TYPE_ALIASES[_alias_key] = _alias_target


def resolve_tumor_type(raw_value: str) -> Tuple[Optional[str], Optional[Dict]]:
    key = re.sub(r"\s+", " ", (raw_value or "").strip()).lower()
    canonical = TUMOR_TYPE_ALIASES.get(key)
    if canonical is None:
        return None, None
    return canonical, TUMOR_TYPE_CONFIG[canonical]


# =====================================================================
# REQUEST CATEGORY (surgery / scan) - Column D, OPTIONAL
# =====================================================================
# Practical way to flag surgical-operation or scan rows without touching
# the tumor type/diagnosis code: an optional 4th Excel column. Blank ->
# ordinary request (tumor_cfg's own proc_id, e.g. 806 breast / 10 other).
# "surgery" -> proc_id forced to 486 regardless of tumor type. "scan" ->
# proc_id forced to 141 regardless of tumor type. diag_code and
# speciality_code always stay exactly what the tumor type (Column C)
# says - only proc_id is affected by this column.
REQUEST_CATEGORY_ALIASES = {
    "": "ordinary",
    "ordinary": "ordinary",
    "normal": "ordinary",
    "surgery": "surgery",
    "surgical": "surgery",
    "operation": "surgery",
    "عملية": "surgery",
    "عمليه": "surgery",
    "جراحة": "surgery",
    "جراحه": "surgery",
    "scan": "scan",
    "scans": "scan",
    "imaging": "scan",
    "radiology": "scan",
    "اشعة": "scan",
    "اشعه": "scan",
    "سكان": "scan",
}

REQUEST_CATEGORY_PROC_ID_OVERRIDE = {
    "surgery": "486",
    "scan": "141",
    # "ordinary" intentionally absent -> falls back to tumor_cfg["proc_id"]
}


def resolve_request_category(raw_value: str) -> Optional[str]:
    """Returns 'ordinary' / 'surgery' / 'scan', or None if the Excel
    value doesn't match any known category (row should then be logged
    FAILED rather than guessed at, same policy as unknown tumor types)."""
    key = re.sub(r"\s+", " ", (raw_value or "").strip()).lower()
    return REQUEST_CATEGORY_ALIASES.get(key)


def resolve_effective_proc_id(tumor_cfg: Dict, request_category: str) -> str:
    """The proc_id actually used for THIS row: tumor_cfg's own proc_id,
    unless request_category is 'surgery' or 'scan', in which case that
    override wins. diag_code/speciality_code are untouched either way."""
    return REQUEST_CATEGORY_PROC_ID_OVERRIDE.get(request_category, tumor_cfg["proc_id"])


# =====================================================================
# DERIVED PATHS & LOGGING
# =====================================================================

RUN_TAG = Path(EXCEL_INPUT_PATH).stem
DEBUG_DIR = os.path.join(SUBMISSION_ROOT, f"debug_{RUN_TAG}")
RUN_LOG_PATH = os.path.join(SUBMISSION_ROOT, f"Submission_Run_Log_{RUN_TAG}.xlsx")
CHECKPOINT_PATH = os.path.join(SUBMISSION_ROOT, f"Row_Checkpoints_{RUN_TAG}.json")

os.makedirs(SUBMISSION_ROOT, exist_ok=True)
os.makedirs(DEBUG_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(os.path.join(SUBMISSION_ROOT, f"unified_run_{RUN_TAG}.log"), encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# --------------------------------------------------------------------
# Medical report PDF generation lives in medical_report_overlay.py.
# It draws the Name / National ID / treatment-plan text directly onto a
# pre-exported PDF of the template (reportlab + PyPDF2, the same overlay
# technique used below for signatures/stamp) - no Word, no COM, no
# subprocess, no shared lock or queue. Fully safe to call from any number
# of parallel copies of this script running at the same time.
# --------------------------------------------------------------------
from medical_report_overlay import build_medical_report_pdf  # noqa: E402

# --------------------------------------------------------------------
# Patient-document sourcing/refresh chain lives in two separate modules
# (see each file's own header for full detail on how each was verified
# against real captured site responses):
#
#   1. patient_pdf_dms_archive_fallback.py - talks to the on-prem
#      CMIS/MRM portal. Used TWO ways in Stage 4 below, run BEFORE the
#      website fallback:
#        a) If a local PDF was already found for this patient, first
#           REFRESH it in place - check CMIS's archive for anything
#           scanned in the last RECENT_ARCHIVE_DAYS_BACK days (default
#           7) and not yet in the local file, merge it on, overwrite
#           the same local file.
#        b) If no local PDF exists at all, do a FULL extraction of
#           every archived PDF for this patient from CMIS and save it
#           as a new file.
#   2. patient_pdf_website_fallback.py - re-searches the SMC decree
#      site itself for a Gustave-facility request with a downloadable
#      PDF attached. Tried LAST, only if step 1 still leaves this row
#      with no patient PDF at all (CMIS lookup failed entirely, or the
#      patient truly has nothing archived there).
# --------------------------------------------------------------------
from patient_pdf_dms_archive_fallback import (  # noqa: E402
    get_all_patient_archive_pdfs_merged,
    refresh_local_pdf_with_recent_archive_docs,
)
from patient_pdf_website_fallback import download_patient_pdf_from_website  # noqa: E402

# Where fallback-extracted patient document PDFs (SMC website / CMIS
# archive) get saved when nothing was found locally. This used to point
# directly at PATIENT_DOCS_ROOT. It now points at the UNDER_PROCESSED
# subfolder instead (per your latest instruction) - so a freshly-pulled,
# not-yet-cleaned PDF is easy to find and review/label on its own, and is
# deliberately NOT picked up by find_patient_id_pdf() (which excludes this
# subfolder) until the labeling server has saved a cleaned copy back into
# PATIENT_DOCS_ROOT itself under the same patient ID.
FALLBACK_PATIENT_DOCS_DIR = PATIENT_DOCS_UNDER_PROCESSED_DIR
os.makedirs(FALLBACK_PATIENT_DOCS_DIR, exist_ok=True)
os.makedirs(PATIENT_DOCS_ROOT, exist_ok=True)


# =====================================================================
# ERROR TYPES
# =====================================================================

class RowProcessingError(Exception):
    """A business-logic failure (bad data, validation error, missing
    token, etc.) - NOT retried; recorded straight into the run log."""


class RowBlacklisted(Exception):
    """Patient is on the SMC blacklist - row skipped, not a failure."""


NETWORK_EXCEPTIONS = (
    requests.exceptions.ConnectionError,
    requests.exceptions.Timeout,
    requests.exceptions.ChunkedEncodingError,
)

# wkhtmltopdf runs as a subprocess outside `requests`, so its own
# transient network failures surface as OSError / pdfkit errors instead.
BROAD_NETWORK_EXCEPTIONS = NETWORK_EXCEPTIONS + (OSError,)


def call_with_reconnect(session: "SMCSession", stage_label: str, func, *args,
                         max_retries: int = NETWORK_MAX_RETRIES,
                         retry_wait_seconds: int = NETWORK_RETRY_WAIT_SECONDS,
                         broad: bool = False, **kwargs):
    """
    Runs func(*args, **kwargs). On a network-level exception, re-logs in
    and retries the SAME call (i.e. the whole stage_* function) from
    scratch, up to max_retries times. Business-logic errors
    (RowProcessingError / RowBlacklisted) and any other exception pass
    straight through untouched - only connectivity blips are retried.
    """
    exceptions_to_catch = BROAD_NETWORK_EXCEPTIONS if broad else NETWORK_EXCEPTIONS
    attempt = 0
    while True:
        try:
            return func(*args, **kwargs)
        except exceptions_to_catch as exc:
            attempt += 1
            log.warning(f"    [{stage_label}] network interruption (attempt {attempt}/{max_retries}): {exc}")
            if attempt > max_retries:
                raise RuntimeError(
                    f"[{stage_label}] gave up after {max_retries} reconnect attempts: {exc}"
                ) from exc
            time.sleep(retry_wait_seconds)
            log.info(f"    [{stage_label}] re-logging in and retrying this step from the beginning …")
            if not session.login():
                raise RuntimeError(f"[{stage_label}] re-login failed after network interruption.")


# =====================================================================
# CHECKPOINT FILE (crash / restart resilience)
# =====================================================================

def load_checkpoints(path: str) -> Dict[str, Dict]:
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as exc:
        log.warning(f"Could not read checkpoint file ({exc}) - starting with no checkpoints.")
        return {}


def save_checkpoint(path: str, checkpoints: Dict[str, Dict], key: str, data: Dict):
    checkpoints[key] = data
    tmp_path = path + ".tmp"
    try:
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(checkpoints, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, path)
    except Exception as exc:
        log.error(f"Could not write checkpoint file: {exc}")


# =====================================================================
# SMC SESSION  (union of both scripts' API surfaces)
# =====================================================================

class SMCSession:
    def __init__(self):
        self.s = requests.Session()
        self.s.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept-Language": "ar,en-US;q=0.9,en;q=0.8",
        })
        self.logged_in = False

    # ------------------------------------------------------------------
    # Login
    # ------------------------------------------------------------------

    def login(self) -> bool:
        log.info("Logging in to SMC …")
        try:
            page = self.s.get(f"{BASE_URL}/smc/Home/Index", timeout=30)
            soup = BeautifulSoup(page.text, "html.parser")
            tok = soup.find("input", {"name": "__RequestVerificationToken"})
            data = {"username": USERNAME, "password": PASSWORD}
            if tok:
                data["__RequestVerificationToken"] = tok["value"]
            resp = self.s.post(f"{BASE_URL}/smc/Home/Index", data=data, timeout=30, allow_redirects=True)
            if ("logout" in resp.text.lower() or "dashboard" in resp.text.lower()
                    or "مرحبا" in resp.text or "تسجيل الخروج" in resp.text):
                self.logged_in = True
                log.info("Login successful.")
                return True
            log.error("Login failed — check username/password.")
            return False
        except Exception as exc:
            log.error(f"Login error: {exc}")
            return False

    def _session_expired(self, r: requests.Response) -> bool:
        return r.status_code == 302 or (
            r.status_code == 200 and "Home/Index" in r.url and "username" in r.text.lower()
        )

    def _get(self, url: str, ajax: bool = False, referer: Optional[str] = None, **kw) -> Optional[requests.Response]:
        short = url.split("?")[0].split("/smc/")[-1]
        headers = {}
        if ajax:
            headers["X-Requested-With"] = "XMLHttpRequest"
            if referer:
                headers["Referer"] = referer
        try:
            r = self.s.get(url, timeout=30, headers=headers or None, **kw)
            log.info(f"    GET  {short}  params={kw.get('params')}  -> {r.status_code}")
            if self._session_expired(r):
                log.warning("Session expired — re-logging in …")
                if self.login():
                    r = self.s.get(url, timeout=30, headers=headers or None, **kw)
                else:
                    return None
            if r.status_code != 200:
                log.error(f"    GET {short} returned {r.status_code}. Body: {r.text[:300]!r}")
                return None
            return r
        except NETWORK_EXCEPTIONS:
            raise
        except Exception as exc:
            log.error(f"GET {url}: {exc}")
            return None

    def _post_ajax(self, url: str, data: dict, **kw) -> Optional[requests.Response]:
        """AJAX-style POST for JSON-returning endpoints."""
        short = url.split("?")[0].split("/smc/")[-1]
        headers = {
            "X-Requested-With": "XMLHttpRequest",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        }
        try:
            r = self.s.post(url, data=data, headers=headers, timeout=60, **kw)
            log.info(f"    POST {short}  -> {r.status_code}  ({len(r.text)} chars)")
            if self._session_expired(r):
                log.warning("Session expired — re-logging in …")
                if self.login():
                    r = self.s.post(url, data=data, headers=headers, timeout=60, **kw)
                else:
                    return None
            if r.status_code != 200:
                log.error(f"    POST {short} returned {r.status_code}. Body: {r.text[:300]!r}")
                return None
            return r
        except NETWORK_EXCEPTIONS:
            raise
        except Exception as exc:
            log.error(f"POST {url}: {exc}")
            return None

    def _post_form_submit(self, url: str, data: dict, referer: str) -> Optional[requests.Response]:
        """Real browser-style page-navigation form POST (no XHR headers)."""
        headers = {
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Origin": BASE_URL,
            "Referer": referer,
        }
        try:
            r = self.s.post(url, data=data, headers=headers, timeout=60, allow_redirects=True)
            log.info(f"    POST(form) {url.split('/smc/')[-1]} -> {r.status_code}  ({len(r.text)} chars)")
            if r.status_code != 200:
                log.error(f"    POST(form) returned {r.status_code}. Body: {r.text[:300]!r}")
                return None
            return r
        except NETWORK_EXCEPTIONS:
            raise
        except Exception as exc:
            log.error(f"POST(form) {url}: {exc}")
            return None

    # ------------------------------------------------------------------
    # MDT CREATION FLOW  (PreRequest/* endpoints)
    # ------------------------------------------------------------------

    def get_create_page_context(self) -> Dict[str, str]:
        r = self._get(f"{BASE_URL}/smc/PreRequest/Create")
        if r is None:
            raise RuntimeError("Could not load PreRequest/Create page.")
        soup = BeautifulSoup(r.text, "html.parser")

        def field_value(name: str) -> str:
            tag = soup.find(attrs={"name": name})
            return tag.get("value", "") if tag else ""

        return {
            "token": field_value("__RequestVerificationToken"),
            "user_id": field_value("UserId"),
            "sending_site_id": field_value("SENDINGSITEID"),
            "treatment_hospital": field_value("treatmentHospital"),
        }

    def search_blacklist(self, ssn: str) -> Dict:
        r = self._post_ajax(f"{BASE_URL}/smc/PreRequest/SearchBlackListSSN", data={"SSN": ssn})
        if r is None:
            return {}
        try:
            return r.json()
        except Exception:
            return {}

    def search_ssn_prerequest(self, ssn: str) -> Optional[Dict]:
        """POST /smc/PreRequest/SearchSSN — used during MDT creation."""
        r = self._post_ajax(f"{BASE_URL}/smc/PreRequest/SearchSSN", data={"SSN": ssn})
        if r is None:
            return None
        try:
            return r.json()
        except Exception:
            log.error(f"PreRequest/SearchSSN returned non-JSON for {ssn}")
            return None

    def get_last_letter_content(self, ssn: str) -> str:
        r = self._post_ajax(f"{BASE_URL}/smc/PreRequest/GetLastLetterContent", data={"SSN": ssn})
        return r.text.strip() if r is not None else ""

    def get_regions_by_city(self, city_id) -> List[Dict]:
        r = self._get(f"{BASE_URL}/smc/Requests/GetRegionsByCityId", params={"cityId": city_id})
        try:
            return r.json() if r is not None else []
        except Exception:
            return []

    def get_diagnosis_data(self, speciality_code: str) -> List[Dict]:
        r = self._get(f"{BASE_URL}/smc/Requests/GetDiagnosisData", params={"specialityCode": speciality_code})
        try:
            return r.json() if r is not None else []
        except Exception:
            return []

    def get_diagnosis_with_med_proc(self, diag_code: str) -> List[Dict]:
        r = self._get(f"{BASE_URL}/smc/Requests/GetDiagnosisWithMedProc", params={"diagName": "", "diagCode": diag_code})
        try:
            return r.json() if r is not None else []
        except Exception:
            return []

    def get_department_by_diagnosis(self, diag_code: str) -> Optional[int]:
        r = self._get(f"{BASE_URL}/smc/Requests/GetDepartmentByDiagnosisId", params={"diagCode": diag_code})
        try:
            data = r.json() if r is not None else {}
            return data.get("department")
        except Exception:
            return None

    def get_treatment_procs_by_diagnose_code(self, diag_code: str) -> List[Dict]:
        r = self._get(f"{BASE_URL}/smc/PreRequest/GetRequestTreatmentProcsByDiagnoseCode", params={"diagCode": diag_code})
        try:
            return r.json() if r is not None else []
        except Exception:
            return []

    def get_treatment_proc_desc(self, proc_id: str) -> Optional[str]:
        r = self._get(f"{BASE_URL}/smc/Requests/GetReqTreatmentProcDesc", params={"procId": proc_id})
        try:
            data = r.json() if r is not None else []
            return data[0]["DESCRIPTION"] if data else None
        except Exception:
            return None

    def create_prerequest(self, payload: Dict[str, str]) -> Tuple[Optional[str], str]:
        r = self._post_form_submit(
            f"{BASE_URL}/smc/PreRequest/Create", data=payload,
            referer=f"{BASE_URL}/smc/PreRequest/Create",
        )
        if r is None:
            return None, ""
        m = re.search(r"تم تسجيل إستمارة رقم\s*</strong>\s*:\s*(\d+)", r.text)
        if not m:
            errors = re.findall(
                r'<span[^>]*class="[^"]*field-validation-error[^"]*"[^>]*>(.*?)</span>',
                r.text, re.S,
            ) + re.findall(
                r'<div[^>]*class="[^"]*validation-summary-errors[^"]*"[^>]*>(.*?)</div>',
                r.text, re.S,
            )
            errors = [re.sub(r"<.*?>", "", e).strip() for e in errors if re.sub(r"<.*?>", "", e).strip()]
            if errors:
                log.error(f"    Server-side validation errors: {errors}")
            elif "username" in r.text.lower() and "password" in r.text.lower():
                log.error("    Response looks like a login page — session was likely rejected.")
        return (m.group(1) if m else None), r.text

    def print_prerequest_url(self, pre_request_id: str) -> str:
        return f"{BASE_URL}/smc/PreRequest/PrintPreRequest?preRequestId={pre_request_id}"

    def cookie_list(self) -> List[Tuple[str, str]]:
        return [(c.name, c.value) for c in self.s.cookies]

    # ------------------------------------------------------------------
    # FINAL UPLOAD FLOW  (Requests/* endpoints)
    # ------------------------------------------------------------------

    def get_pre_requests(self, pre_request_id: str) -> Optional[dict]:
        """Best-effort listing lookup — a failure here is non-fatal."""
        params = {
            "preRequestId": pre_request_id, "nationalId": "",
            "fromDate": (datetime.now().replace(day=1)).strftime("%m-%d-%Y"),
            "toDate": datetime.now().strftime("%m-%d-%Y"),
            "page": "1",
        }
        r = self._get(f"{BASE_URL}/smc/PreRequest/GetPreRequests", params=params)
        if r is None:
            return None
        try:
            return r.json()
        except Exception:
            return {}

    def get_requests_create_context(self, pre_request_id: str) -> Dict[str, str]:
        url = f"{BASE_URL}/smc/Requests/Create?action=HopitalCreateNewRequest&prid={pre_request_id}"
        r = self._get(url)
        if r is None:
            raise RuntimeError(f"Could not load Requests/Create page for prid={pre_request_id}")

        soup = BeautifulSoup(r.text, "html.parser")

        def field_val(name: str) -> str:
            tag = soup.find(attrs={"name": name}) or soup.find(attrs={"id": name})
            return (tag.get("value") or "").strip() if tag else ""

        ctx = {
            "token": field_val("__RequestVerificationToken"),
            "user_id": field_val("UserId"),
            "prerequestidfk": field_val("PREREQUESTIDFK") or pre_request_id,
        }
        if not ctx["token"]:
            raise RuntimeError(
                f"Could not scrape __RequestVerificationToken from Requests/Create page for prid={pre_request_id}"
            )
        return ctx

    def search_ssn_requests(self, patient_id: str, pre_request_id: str) -> Optional[dict]:
        """POST /smc/Requests/SearchSSN — used during final upload (different
        endpoint from search_ssn_prerequest, deliberately named apart)."""
        r = self._post_ajax(
            f"{BASE_URL}/smc/Requests/SearchSSN",
            data={"SSN": patient_id, "checkDecree": "false", "preRequestID": pre_request_id},
        )
        if r is None:
            return None
        try:
            return r.json()
        except Exception:
            log.error(f"Requests/SearchSSN returned non-JSON: {r.text[:300]!r}")
            return None

    def get_file_size(self, referer: Optional[str] = None):
        self._get(f"{BASE_URL}/smc//Requests/GetFileSize", ajax=True, referer=referer)

    def get_sms_setup(self, referer: Optional[str] = None):
        self._get(f"{BASE_URL}/smc//Requests/GetSMSSetup", ajax=True, referer=referer)

    def submit_request_with_pdf(self, form_fields: dict, pdf_path: str, referer_prid: str) -> Tuple[Optional[str], str]:
        """POST /smc/Requests/Create as a real multipart browser form,
        carrying the merged PDF in 'medicalReportImg'. Returns
        (new_request_number, raw_html)."""
        url = f"{BASE_URL}/smc/Requests/Create"
        empty_file = ("", b"", "application/octet-stream")
        pdf_filename = os.path.basename(pdf_path)

        try:
            with open(pdf_path, "rb") as fh:
                pdf_bytes = fh.read()
        except OSError as exc:
            log.error(f"Cannot read PDF {pdf_path}: {exc}")
            return None, ""

        files = [
            ("medicalReportImg", (pdf_filename, pdf_bytes, "application/pdf")),
            ("ssnCardImg", empty_file),
            ("InusranceSiteFile", empty_file),
            ("otherFiles", empty_file),
        ]
        headers = {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Origin": BASE_URL,
            "Referer": f"{BASE_URL}/smc/Requests/Create?action=HopitalCreateNewRequest&prid={referer_prid}",
        }

        try:
            r = self.s.post(url, data=form_fields, files=files, headers=headers, timeout=120, allow_redirects=True)
            log.info(f"    POST(multipart) Requests/Create -> {r.status_code}  ({len(r.text)} chars)")
            if self._session_expired(r):
                log.warning("Session expired on final submit — cannot auto-retry the upload transparently.")
                return None, r.text
            if r.status_code != 200:
                log.error(f"    Requests/Create returned {r.status_code}. Body: {r.text[:300]!r}")
                return None, r.text

            m = re.search(r"رقم الطلب\s*</strong>\s*:\s*(\d+)", r.text)
            if m:
                return m.group(1), r.text
            m2 = re.search(r"رقم\s*الطلب\s*[:\s]+(\d+)", r.text)
            if m2:
                return m2.group(1), r.text
            return None, r.text
        except NETWORK_EXCEPTIONS:
            raise
        except Exception as exc:
            log.error(f"POST(multipart) Requests/Create: {exc}")
            return None, ""


# =====================================================================
# HTML -> PDF  (wkhtmltopdf via pdfkit)
# =====================================================================

def render_print_page_to_pdf(session: SMCSession, pre_request_id: str) -> bytes:
    """Render the authenticated MDT print page to PDF.

    FIXED BUG (superseding the previous cookie-forwarding approach): having
    wkhtmltopdf fetch the URL itself — even with every cookie manually
    forwarded via --cookie, even with a patched-Qt build — proved
    unreliable in this environment: sometimes it silently got the
    logged-out landing page (exit 0, no error), and on retry it sometimes
    hung indefinitely waiting on the network instead of failing. wkhtmltopdf
    doing its own HTTP fetch means its own auth/cookie/header handling has
    to exactly match the site's expectations, with no visibility into why
    it doesn't.

    The `requests.Session` used for every other call in this script is
    already authenticated and has never had this problem. So: fetch the
    print page's HTML ourselves with that trusted session, validate it's
    actually the print form (not a login/landing page) BEFORE spending any
    time on wkhtmltopdf, then hand wkhtmltopdf the HTML directly via stdin.
    wkhtmltopdf then does no networking or auth of its own for the main
    page at all — only for any referenced sub-resources (rare on a
    server-rendered print page), which is a much smaller failure surface
    than the whole page fetch.
    """
    url = session.print_prerequest_url(pre_request_id)

    resp = session.s.get(url, timeout=30)
    if resp.status_code != 200:
        raise OSError(
            f"Could not fetch MDT print page for pre_request_id={pre_request_id} "
            f"(HTTP {resp.status_code}). Will re-login and retry."
        )

    html = resp.text
    looks_like_login_page = ("اسم المستخدم" in html and "كلمة السر" in html)
    looks_like_print_form = ("طلب علاج" in html or "تقرير اللجنة الثلاثية" in html
                              or str(pre_request_id) in html)
    if looks_like_login_page or not looks_like_print_form:
        raise OSError(
            f"Fetched MDT print page for pre_request_id={pre_request_id} looks like the "
            f"logged-out SMC landing/login page, not the actual print form — session was not "
            f"authenticated for this request. Will re-login and retry."
        )

    # wkhtmltopdf is no longer fetching the page, so relative asset URLs
    # (css/images referenced without a full domain) need an explicit base
    # to still resolve correctly.
    base_tag = f'<base href="{BASE_URL}/">'
    if re.search(r"<head[^>]*>", html, re.I):
        html = re.sub(r"(<head[^>]*>)", r"\1" + base_tag, html, count=1, flags=re.I)
    else:
        html = base_tag + html

    cmd = [
        WKHTMLTOPDF_PATH,
        "--encoding", "UTF-8",
        "--quiet",
        "--page-size", "A4",
        "--margin-top", "5mm",
        "--margin-bottom", "5mm",
        "--margin-left", "5mm",
        "--margin-right", "5mm",
        "--no-outline",
        "--load-error-handling", "ignore",
        "--load-media-error-handling", "ignore",
        "-", "-",  # "-" "-" = read HTML from stdin, write PDF to stdout
    ]

    try:
        proc = subprocess.run(
            cmd,
            input=html.encode("utf-8"),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=RENDER_TIMEOUT_SECONDS,
        )
    except subprocess.TimeoutExpired as exc:
        raise OSError(
            f"wkhtmltopdf timed out rendering pre_request_id={pre_request_id} even from "
            f"already-fetched local HTML (likely a hung external asset request) — will retry."
        ) from exc

    if proc.returncode != 0 or not proc.stdout:
        stderr_text = (proc.stderr or b"").decode("utf-8", errors="replace").strip()
        raise RuntimeError(
            "wkhtmltopdf failed: "
            + (stderr_text or f"exit code {proc.returncode}")
        )

    return proc.stdout

    return proc.stdout


# =====================================================================
# SIGNATURE / STAMP OVERLAY
# =====================================================================

def make_signature_blue(image_path: str) -> Image.Image:
    img = Image.open(image_path).convert("RGBA")
    alpha = img.getchannel("A")
    alpha = ImageEnhance.Contrast(alpha).enhance(2.0)
    alpha = ImageEnhance.Brightness(alpha).enhance(1.8)
    new_pixels = []
    for (_, _, _, a), new_a in zip(img.getdata(), alpha.getdata()):
        if new_a > 10:
            new_pixels.append((20, 60, 180, min(255, int(new_a * 1.2))))
        else:
            new_pixels.append((255, 255, 255, 0))
    img.putdata(new_pixels)
    return img


def process_stamp(stamp_path: str) -> Image.Image:
    img = Image.open(stamp_path).convert("RGBA")
    alpha = img.getchannel("A")
    alpha = ImageEnhance.Contrast(alpha).enhance(3.0)
    alpha = ImageEnhance.Brightness(alpha).enhance(2.5)
    threshold = 30
    alpha_binary = alpha.point(lambda p: 255 if p > threshold else 0).convert("L")
    for _ in range(2):
        alpha_binary = alpha_binary.filter(ImageFilter.MaxFilter(3))
    alpha_binary = alpha_binary.filter(ImageFilter.GaussianBlur(radius=1))
    alpha_binary = ImageEnhance.Contrast(alpha_binary).enhance(3.0)
    alpha_binary = ImageEnhance.Brightness(alpha_binary).enhance(1.5)

    new_pixels = []
    for alpha_value in alpha_binary.getdata():
        if alpha_value > 10:
            new_pixels.append((30, 70, 210, min(255, int(alpha_value * 1.3))))
        else:
            new_pixels.append((255, 255, 255, 0))
    final_img = Image.new("RGBA", img.size)
    final_img.putdata(new_pixels)
    final_img = ImageEnhance.Contrast(final_img).enhance(1.5)
    final_img = ImageEnhance.Sharpness(final_img).enhance(2.0)
    return final_img


def apply_signatures_and_stamp(mdt_pdf_bytes: bytes) -> bytes:
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        processed = []
        for cfg in SIGNATURES:
            src = SIGNATURE_FILES[cfg["file_key"]]
            if not os.path.exists(src):
                raise FileNotFoundError(f"Signature file not found: {src}")
            img = make_signature_blue(src)
            path = tmp / f"{cfg['file_key']}.png"
            img.save(path, "PNG")
            processed.append({**cfg, "processed_path": str(path)})

        stamp_src = SIGNATURE_FILES["stamp"]
        if not os.path.exists(stamp_src):
            raise FileNotFoundError(f"Stamp file not found: {stamp_src}")
        stamp_img = process_stamp(stamp_src)
        stamp_path = tmp / "stamp.png"
        stamp_img.save(stamp_path, "PNG")

        packet = io.BytesIO()
        can = canvas.Canvas(packet, pagesize=A4)
        for sig in processed:
            can.drawImage(sig["processed_path"], sig["x"], sig["y"], width=sig["width"],
                           height=sig["height"], mask="auto", preserveAspectRatio=True, anchor="sw")
        can.drawImage(str(stamp_path), STAMP["x"], STAMP["y"], width=STAMP["width"],
                       height=STAMP["height"], mask="auto", preserveAspectRatio=True, anchor="sw")
        can.save()
        packet.seek(0)

        overlay_pdf = PdfReader(packet)
        original_pdf = PdfReader(io.BytesIO(mdt_pdf_bytes))
        output = PdfWriter()

        page = original_pdf.pages[0]
        page.merge_page(overlay_pdf.pages[0])
        output.add_page(page)
        for i in range(1, len(original_pdf.pages)):
            output.add_page(original_pdf.pages[i])

        out_buf = io.BytesIO()
        output.write(out_buf)
        return out_buf.getvalue()


# =====================================================================
# MEDICAL REPORT GENERATION
# =====================================================================
# build_medical_report_pdf() is imported from medical_report_overlay.py
# (see the import near the top of this file). It draws Name / National
# ID / treatment-plan text straight onto a pre-exported PDF of the
# template - no Word, no COM, no queue. See that file's header comment
# for the one-time setup (export template to PDF, calibrate coordinates).

# =====================================================================
# PATIENT DOCUMENT LOOKUP
# =====================================================================

def find_patient_id_pdf(patient_id: str) -> Optional[str]:
    """
    Searches PATIENT_DOCS_ROOT for an already-cleaned PDF named after this
    patient. Deliberately EXCLUDES the UNDER_PROCESSED subfolder: files
    sitting there are freshly extracted but not yet reviewed/cleaned by
    the labeling server, so they must never be silently picked up as if
    they were an already-cleaned local file - that would skip both the
    labeling step and the global continue/skip review entirely.
    """
    candidates = glob.glob(os.path.join(PATIENT_DOCS_ROOT, "**", f"*{patient_id}*.pdf"), recursive=True)
    under_processed_norm = os.path.normcase(os.path.normpath(PATIENT_DOCS_UNDER_PROCESSED_DIR)) + os.sep
    candidates = [
        c for c in candidates
        if not os.path.normcase(os.path.normpath(c)).startswith(under_processed_norm)
    ]
    return candidates[0] if candidates else None


# =====================================================================
# FINAL MERGE
# =====================================================================

def merge_final_pdf(mdt_pdf_bytes: bytes, report_pdf_bytes: bytes, patient_id_pdf_path: str, output_path: str):
    """Order: signed MDT form -> medical report -> patient's full document."""
    merger = PdfMerger()
    merger.append(io.BytesIO(mdt_pdf_bytes))
    merger.append(io.BytesIO(report_pdf_bytes))
    merger.append(patient_id_pdf_path)
    merger.write(output_path)
    merger.close()


# =====================================================================
# SHARED HELPERS
# =====================================================================

def birthdate_from_national_id(national_id: str) -> Optional[str]:
    """
    Egyptian national ID encodes birthdate in its first 7 digits:
      digit 1   : century flag  2 = 1900s, 3 = 2000s
      digits 2-3: two-digit year
      digits 4-5: month
      digits 6-7: day
    """
    if not national_id or not national_id.isdigit() or len(national_id) < 7:
        return None
    century = {"2": "19", "3": "20"}.get(national_id[0])
    if century is None:
        return None
    yy, mm, dd = national_id[1:3], national_id[3:5], national_id[5:7]
    try:
        datetime(int(century + yy), int(mm), int(dd))
    except ValueError:
        return None
    return f"{century}{yy}-{mm}-{dd}"


def parse_committee_date(ms_date: Optional[str]) -> str:
    """Converts a .NET /Date(ms)/ string to YYYY-MM-DD, correcting for
    the site encoding Cairo local midnight rather than UTC midnight."""
    if ms_date:
        m = re.search(r"/Date\((-?\d+)\)/", ms_date)
        if m:
            try:
                ts = int(m.group(1)) / 1000
                utc_dt = datetime.utcfromtimestamp(ts)
                local_dt = utc_dt + timedelta(hours=CAIRO_UTC_OFFSET_HOURS)
                return local_dt.strftime("%Y-%m-%d")
            except Exception:
                pass
    return datetime.now().strftime("%Y-%m-%d")


def dump_debug_artifacts(patient_id: str, payload: Dict[str, str], raw_html: str):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    payload_path = os.path.join(DEBUG_DIR, f"{patient_id}_{ts}_mdt_payload.json")
    html_path = os.path.join(DEBUG_DIR, f"{patient_id}_{ts}_mdt_response.html")
    try:
        with open(payload_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(raw_html)
        log.info(f"    Debug artifacts written: {payload_path} , {html_path}")
    except Exception as exc:
        log.error(f"    Could not write debug artifacts: {exc}")


def save_debug_html(patient_id: str, pre_request_id: str, html: str):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(DEBUG_DIR, f"{patient_id}_{pre_request_id}_{ts}_upload_response.html")
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)
        log.info(f"    Debug HTML saved: {path}")
    except Exception as exc:
        log.error(f"    Could not save debug HTML: {exc}")


# =====================================================================
# STAGE 1: MDT CREATION
# =====================================================================

def stage_create_mdt(session: SMCSession, patient_id: str, description: str, tumor_cfg: Dict) -> Dict:
    """
    Runs the full MDT-creation sequence. Safe to re-run from scratch if
    interrupted partway through - nothing here is persisted server-side
    until create_prerequest() actually succeeds.
    """
    warnings: List[str] = []

    bl = session.search_blacklist(patient_id)
    if bl.get("exist"):
        raise RowBlacklisted(f"Blacklist check returned: {bl}")

    patient_resp = session.search_ssn_prerequest(patient_id)
    if not patient_resp or not patient_resp.get("patient"):
        raise RowProcessingError(f"PreRequest/SearchSSN returned no patient data. Raw: {str(patient_resp)[:300]}")
    patient = patient_resp["patient"]
    city_id = patient_resp.get("cityId")
    hio_data_id = patient_resp.get("HIODATAID") or patient_resp.get("HIODataID")
    insurance_no = patient_resp.get("insuranceNo")

    full_name = re.sub(r"\s+", " ", (patient.get("CITIZENFULLNAMEARABIC") or "")).strip()
    if not full_name:
        parts = [patient.get(k) or "" for k in
                 ("CITIZENFIRSTNAMEARABIC", "CITIZENMIDDNAMEARABIC",
                  "CITIZENLASTNAMEARABIC", "CITIZENFAMILYNAMEARABIC")]
        full_name = re.sub(r"\s+", " ", " ".join(parts)).strip()
    if not full_name:
        raise RowProcessingError("Could not determine patient full name from SearchSSN response.")

    last_letter = session.get_last_letter_content(patient_id)
    if last_letter not in ('""', ""):
        warnings.append(f"GetLastLetterContent returned non-empty: {last_letter[:200]}")

    session.get_regions_by_city(city_id)
    session.get_diagnosis_data(tumor_cfg["speciality_code"])
    diag_matches = session.get_diagnosis_with_med_proc(tumor_cfg["diag_code"])
    initial_icd10_name = diag_matches[0]["DIAGNOSISARABICNAME"] if diag_matches else ""

    department_id = session.get_department_by_diagnosis(tumor_cfg["diag_code"])
    if not department_id:
        warnings.append("GetDepartmentByDiagnosisId returned no department; proceeding without it.")

    proc_list = session.get_treatment_procs_by_diagnose_code(tumor_cfg["diag_code"])
    if not any(str(p.get("ID")) == tumor_cfg["proc_id"] for p in proc_list):
        warnings.append(f"TREATMENTPROCEDUREID {tumor_cfg['proc_id']} not found in "
                         f"GetRequestTreatmentProcsByDiagnoseCode results.")

    proc_desc = session.get_treatment_proc_desc(tumor_cfg["proc_id"])
    if not proc_desc:
        warnings.append("GetReqTreatmentProcDesc returned no description.")

    try:
        ctx = session.get_create_page_context()
    except RuntimeError as exc:
        raise RowProcessingError(str(exc)) from exc
    if not ctx.get("token"):
        raise RowProcessingError("Could not scrape __RequestVerificationToken from PreRequest/Create page.")
    if not ctx.get("user_id") or not ctx.get("sending_site_id"):
        warnings.append("UserId or SENDINGSITEID came back empty from the Create page — "
                         "likely cause if creation fails.")

    birthdate_str = birthdate_from_national_id(patient_id) or ""
    if not birthdate_str:
        warnings.append(f"Could not derive a valid birthdate from national ID {patient_id!r}.")

    gender_code = GENDER_MAP.get(patient.get("GENDER"), "")
    if not gender_code:
        warnings.append(f"Unmapped GENDER value: {patient.get('GENDER')!r}")

    has_insurance = "Y" if insurance_no else "N"

    payload = {
        "__RequestVerificationToken": ctx["token"],
        "ACTIVE": "Y",
        "CITIZEN.ACTIVE": "Y",
        "UserId": ctx["user_id"],
        "HIODataID": str(hio_data_id or ""),
        "CITIZENSSN": patient_id,
        "CITIZEN.DEATHDATE": "",
        "CITIZEN.CITIZENFIRSTNAMEARABIC": patient.get("CITIZENFIRSTNAMEARABIC") or "",
        "CITIZEN.CITIZENMIDDNAMEARABIC": patient.get("CITIZENMIDDNAMEARABIC") or "",
        "CITIZEN.CITIZENLASTNAMEARABIC": patient.get("CITIZENLASTNAMEARABIC") or "",
        "CITIZEN.CITIZENFAMILYNAMEARABIC": patient.get("CITIZENFAMILYNAMEARABIC") or "",
        "CITIZEN.CITIZENBIRTHDATE": birthdate_str,
        "CITIZEN.MARITALID": str(patient.get("MARITALID") or ""),
        "CITIZEN.GENDER": gender_code,
        "CITIZEN.CITYOFBIRTH": str(city_id or ""),
        "CITYOFBIRTH": str(city_id or ""),
        "REGIONID": str(patient.get("REGIONID") or ""),
        "ADDRESS": patient.get("ADDRESS") or "",
        "CITIZEN.CITIZENPHONE": patient.get("CITIZENPHONE") or "",
        "CITIZEN.CITIZENMOBILE": patient.get("CITIZENMOBILE") or "",
        "CITIZEN.PROFESSIONID": str(patient.get("PROFESSIONID") or ""),
        "treatmentHospital": ctx["treatment_hospital"],
        "SENDINGSITEID": ctx["sending_site_id"],
        "RecommHospitalName": ctx["treatment_hospital"],
        "RECOMMHOSPITALID": ctx["sending_site_id"],
        "sendingHospital": ctx["treatment_hospital"],
        "SENDINGHOSPITALID": ctx["sending_site_id"],
        "FromView": "CreateRequest",
        "RecExternalSite": "1",
        "HASINSURANCE": has_insurance,
        "INSURANCESITEIDFK": "",
        "REQUESTIMPORTANCEID": REQUEST_IMPORTANCE_ID,
        "COMMITTEEDATE": datetime.now().strftime("%Y-%m-%d"),
        "initialICD10Name": initial_icd10_name,
        "INITIALICD10CODE": tumor_cfg["diag_code"],
        "TREATMENTPROCEDUREID": tumor_cfg["proc_id"],
        "RENALLFAILURESESSIONSTARTDATE": "",
        "TREATMENTPLAN": description,
        "DIAGNOSISGROUP": tumor_cfg["speciality_code"],
        "DEPARTMENTIDFK": str(department_id or ""),
        "INITIALICD10CODECOMMENT": "",
        "SPECIALCOMMITTEEDOCTOR1": SPECIALCOMMITTEEDOCTOR1,
        "SPECIALCOMMITTEEDOCTOR2": SPECIALCOMMITTEEDOCTOR2,
        "SPECIALCOMMITTEEDOCTOR3": SPECIALCOMMITTEEDOCTOR3,
        "ICUREQUESTID": "", "ICUFROMDATE": "", "ICUTODATE": "", "ICUHOSPID": "",
        "ICUHOSPNAME": "", "ICUSTATUSID": "", "ICUSTATUSNAME": "",
        "ICUREQUESTTYPEID": "", "ICUREQUESTTYPENAME": "", "ICUSERVICEID": "",
        "ICUSERVICENAME": "", "ICUSERVICETYPEID": "", "ICUSERVICETYPENAME": "",
    }

    pre_request_id, raw_html = session.create_prerequest(payload)
    if not pre_request_id:
        dump_debug_artifacts(patient_id, payload, raw_html)
        raise RowProcessingError(
            "Create did not return a request number — creation likely failed. "
            f"Raw response + payload dumped to {DEBUG_DIR}\\{patient_id}_* for inspection."
        )

    return {"pre_request_id": pre_request_id, "full_name": full_name, "warnings": warnings}


# =====================================================================
# STAGE 2: RENDER + SIGN MDT FORM
# =====================================================================

def stage_render_and_sign(session: SMCSession, pre_request_id: str) -> bytes:
    mdt_pdf_bytes = render_print_page_to_pdf(session, pre_request_id)
    return apply_signatures_and_stamp(mdt_pdf_bytes)


# =====================================================================
# STAGE 6: FINAL UPLOAD
# =====================================================================

def stage_upload_merged_pdf(session: SMCSession, patient_id: str, pre_request_id: str, merged_pdf_path: str,
                             tumor_cfg: Optional[Dict] = None) -> str:
    pr_data = session.get_pre_requests(pre_request_id)
    if pr_data is None:
        log.warning("    GetPreRequests best-effort call failed — proceeding anyway.")

    try:
        ctx = session.get_requests_create_context(pre_request_id)
    except RuntimeError as exc:
        raise RowProcessingError(str(exc)) from exc

    create_page_referer = f"{BASE_URL}/smc/Requests/Create?action=HopitalCreateNewRequest&prid={pre_request_id}"

    ssn_resp = session.search_ssn_requests(patient_id, pre_request_id)
    if not ssn_resp or not ssn_resp.get("patient"):
        raise RowProcessingError(f"Requests/SearchSSN returned no patient data. Raw: {str(ssn_resp)[:300]}")

    patient = ssn_resp["patient"]
    pre_req = ssn_resp.get("patientPreReq") or {}
    insurance_no = ssn_resp.get("insuranceNo") or ""

    birthdate = birthdate_from_national_id(patient_id) or ""
    gender_code = GENDER_MAP.get(patient.get("GENDER", ""), "")
    has_insurance = "Y" if insurance_no else "N"
    committee_date = parse_committee_date(pre_req.get("COMMITTEEDATE"))
    hospital_id = str(pre_req.get("RECOMMHOSPITALID") or "")
    recomend_hospital_name = (pre_req.get("RecommHospitalName") or "").strip()
    sending_hospital = (pre_req.get("sendingHospital") or recomend_hospital_name).strip()

    session.get_file_size(referer=create_page_referer)
    session.get_sms_setup(referer=create_page_referer)

    form_fields = {
        "__RequestVerificationToken": ctx["token"],
        "CANREQSTATUSID": "0",
        "ACTIVE": "Y",
        "CITIZEN.ACTIVE": "Y",
        "SENDSMS": "",
        "UserId": ctx["user_id"],
        "PREREQUESTIDFK": ctx["prerequestidfk"],
        "PatientSSN": patient_id,
        "CITIZENSSN": patient_id,
        "CITIZEN.DEATHDATE": "",
        "CITIZEN.CITIZENFIRSTNAMEARABIC": patient.get("CITIZENFIRSTNAMEARABIC") or "",
        "CITIZEN.CITIZENMIDDNAMEARABIC": patient.get("CITIZENMIDDNAMEARABIC") or "",
        "CITIZEN.CITIZENLASTNAMEARABIC": patient.get("CITIZENLASTNAMEARABIC") or "",
        "CITIZEN.CITIZENFAMILYNAMEARABIC": patient.get("CITIZENFAMILYNAMEARABIC") or "",
        "CITIZEN.CITIZENBIRTHDATE": birthdate,
        "CITIZEN.MARITALID": str(patient.get("MARITALID") or ""),
        "CITIZEN.GENDER": gender_code,
        "CITIZEN.CITYOFBIRTH": str(patient.get("CITYOFBIRTH") or ""),
        "REGIONID": str(patient.get("REGIONID") or ""),
        "ADDRESS": patient.get("ADDRESS") or "",
        "CITIZEN.CITIZENPHONE": patient.get("CITIZENPHONE") or "",
        "CITIZEN.CITIZENMOBILE": patient.get("CITIZENMOBILE") or "",
        "CITIZEN.PROFESSIONID": str(patient.get("PROFESSIONID") or ""),
        "treatmentHospital": recomend_hospital_name,
        "SENDINGSITEID": hospital_id,
        "RecommHospitalName": recomend_hospital_name,
        "RECOMMHOSPITALID": hospital_id,
        "sendingHospital": sending_hospital,
        "SENDINGHOSPITALID": str(pre_req.get("SENDINGHOSPITALID") or hospital_id),
        "FromView": "CreateRequest",
        "RecExternalSite": "1",
        "HASINSURANCE": has_insurance,
        "INSURANCESITEIDFK": "",
        "REQUESTIMPORTANCEID": str(pre_req.get("REQUESTIMPORTANCEID") or "22"),
        "COMMITTEEDATE": committee_date,
        "initialICD10Name": pre_req.get("initialICD10Name") or "",
        "INITIALICD10CODE": pre_req.get("INITIALICD10CODE") or "",
        "TREATMENTPROCEDUREID": str(pre_req.get("TREATMENTPROCEDUREID") or ""),
        "RENALFAILURESESSIONSTARTDATE": "",
        "INITIALICD10CODECOMMENT": pre_req.get("INITIALICD10CODECOMMENT") or "",
        # FIXED BUG: this was hardcoded to "2" regardless of tumor type,
        # which silently overwrote the correct DIAGNOSISGROUP (speciality
        # code, e.g. "11" for breast) that was actually used when the MDT
        # was created — meaning the final uploaded request could carry a
        # different committee/speciality routing than the MDT record it's
        # attached to. Now reuses whatever GetPreRequests/SearchSSN
        # actually returned for this pre-request (the value the server
        # itself stored at creation time), falling back to the tumor
        # type's own speciality_code only if the server didn't return one.
        "DIAGNOSISGROUP": str(pre_req.get("DIAGNOSISGROUP") or (tumor_cfg or {}).get("speciality_code") or "9"),
        "DEPARTMENTIDFK": str(pre_req.get("DEPARTMENTIDFK") or ""),
        "TREATMENTPLAN": pre_req.get("TREATMENTPLAN") or "",
        "CANCERPLACEID": "",
        "CANCERDESCID": "",
        "PATIENTCOMPLAIN": "",
    }

    new_req_no, raw_html = session.submit_request_with_pdf(form_fields, merged_pdf_path, referer_prid=pre_request_id)

    if new_req_no:
        return new_req_no

    errors = re.findall(
        r'<span[^>]*class="[^"]*field-validation-error[^"]*"[^>]*>(.*?)</span>',
        raw_html, re.S,
    ) + re.findall(
        r'<div[^>]*class="[^"]*validation-summary-errors[^"]*"[^>]*>(.*?)</div>',
        raw_html, re.S,
    )
    errors = [re.sub(r"<.*?>", "", e).strip() for e in errors if re.sub(r"<.*?>", "", e).strip()]
    save_debug_html(patient_id, pre_request_id, raw_html)
    if errors:
        raise RowProcessingError("Upload validation errors: " + " | ".join(errors))
    if "username" in raw_html.lower() and "password" in raw_html.lower():
        raise RowProcessingError("Upload response looks like a login page — session was rejected.")
    raise RowProcessingError(
        f"Upload POST returned 200 but no 'رقم الطلب' found in response. Debug HTML saved to {DEBUG_DIR}"
    )


# =====================================================================
# ROW PROCESSING (orchestrates all 6 stages, with checkpointing)
# =====================================================================

# =====================================================================
# STAGE 4 HELPER — patient document PDF location, with the flipped
# fallback order + "newly extracted" flag, factored out into its own
# function so it can be unit-tested offline (see run_self_test() near
# the bottom of this file) by passing in fake versions of every
# network-touching argument instead of the real ones.
# =====================================================================

def locate_patient_document_pdf(
    session: "SMCSession",
    patient_id: str,
    *,
    find_local_fn=None,
    website_fn=None,
    refresh_fn=None,
    full_archive_fn=None,
    reconnect_fn=None,
) -> Tuple[Optional[str], Optional[str], bool]:
    """
    Resolves the patient document PDF to use for this row, per the
    (flipped) order you specified:

        1. Local folder (PATIENT_DOCS_ROOT).
             FOUND  -> always refresh with the last
                       RECENT_ARCHIVE_DAYS_BACK days (default 7) of
                       CMIS archive pages first (merged + overwritten
                       in place), then use it. NOT "newly extracted".
        2. SMC website (patient_pdf_website_fallback.py) — tried FIRST
           among the two fallbacks now, only reached if step 1 found
           nothing.
             FOUND  -> also check the CMIS archive for the same
                       RECENT_ARCHIVE_DAYS_BACK-day window and merge
                       that onto the freshly-downloaded file before
                       it's used. "newly extracted" = True.
        3. CMIS full archive extraction (patient_pdf_dms_archive_
           fallback.py) — final fallback, only reached if step 2 found
           nothing either.
             FOUND  -> "newly extracted" = True.
        4. Nothing anywhere -> (None, None, False).

    Every dependency is injectable (defaults to the real production
    functions/module-level objects) purely so run_self_test() can swap
    in fakes and exercise every branch without a network connection,
    real credentials, or real files.

    Returns (pdf_path_or_None, source_or_None, newly_extracted).
    source is one of: "local", "website+archive", "cmis_full_archive".

    Saves into FALLBACK_PATIENT_DOCS_DIR (== PATIENT_DOCS_ROOT), which
    is exactly the folder find_patient_id_pdf() searches — so a
    freshly-extracted file is found LOCALLY on every subsequent run
    instead of being re-extracted from SMC/CMIS every time.
    """
    find_local_fn = find_local_fn or find_patient_id_pdf
    website_fn = website_fn or download_patient_pdf_from_website
    refresh_fn = refresh_fn or refresh_local_pdf_with_recent_archive_docs
    full_archive_fn = full_archive_fn or get_all_patient_archive_pdfs_merged
    reconnect_fn = reconnect_fn or call_with_reconnect

    # ---- Step 1: local folder ----
    id_pdf_path = find_local_fn(patient_id)
    if id_pdf_path:
        log.info(f"  Found local file {id_pdf_path} - checking CMIS archive for newer pages "
                 f"(last {RECENT_ARCHIVE_DAYS_BACK} day(s)) …")
        try:
            refreshed = refresh_fn(patient_id, id_pdf_path, days_back=RECENT_ARCHIVE_DAYS_BACK)
            if refreshed:
                log.info("  Local file updated with newly archived page(s).")
            else:
                log.info("  Nothing new in the archive - using local file as-is.")
        except Exception as exc:
            log.warning(f"  Archive refresh check raised an error (using local file as-is): {exc}")
        return id_pdf_path, "local", False

    # ---- Step 2: SMC website (tried FIRST of the two fallbacks) ----
    log.info(f"  Not found under {PATIENT_DOCS_ROOT} - trying SMC website first …")
    try:
        id_pdf_path = reconnect_fn(
            session, "website PDF fallback", website_fn,
            session, BASE_URL, patient_id, FALLBACK_PATIENT_DOCS_DIR, broad=True,
        )
    except Exception as exc:
        log.warning(f"  Website fallback raised an error, treating as not-found: {exc}")
        id_pdf_path = None

    if id_pdf_path:
        log.info(f"  Found on SMC website - checking CMIS archive for the last "
                 f"{RECENT_ARCHIVE_DAYS_BACK} day(s) of docs to merge in …")
        try:
            refreshed = refresh_fn(patient_id, id_pdf_path, days_back=RECENT_ARCHIVE_DAYS_BACK)
            if refreshed:
                log.info("  Website PDF updated with newly archived CMIS page(s).")
            else:
                log.info(f"  Nothing new in the CMIS archive for the last "
                         f"{RECENT_ARCHIVE_DAYS_BACK} day(s).")
        except Exception as exc:
            log.warning(f"  CMIS archive merge-in raised an error (using website PDF as-is): {exc}")
        return id_pdf_path, "website+archive", True

    # ---- Step 3: CMIS full archive extraction (final fallback) ----
    log.info("  Not on the SMC website either - trying full CMIS archive extraction "
             "(final fallback) …")
    try:
        id_pdf_path = full_archive_fn(patient_id, FALLBACK_PATIENT_DOCS_DIR)
    except Exception as exc:
        log.warning(f"  CMIS archive extraction raised an error, treating as not-found: {exc}")
        id_pdf_path = None

    if id_pdf_path:
        return id_pdf_path, "cmis_full_archive", True

    # ---- Step 4: nothing anywhere ----
    return None, None, False


_DOC_SOURCE_LABELS = {
    "local": "your local patient-documents folder",
    "website+archive": "the SMC website (merged with any recent-week CMIS archive pages)",
    "cmis_full_archive": "the CMIS archive — full extraction (SMC website had nothing)",
}


def prompt_review_extracted_pdf(patient_id: str, pdf_path: str, doc_source: str,
                                 pre_request_id: str, input_fn=input) -> str:
    """
    The "important pause" you asked for: called ONLY when the patient
    document PDF was NOT already sitting locally and had to be pulled
    by one of the fallbacks (doc_source in {"website+archive",
    "cmis_full_archive"}). Never called for doc_source == "local".

    Blocks until the user answers, then returns "continue" or "skip".
    input_fn is injectable so run_self_test() can simulate keystrokes
    without a real terminal.
    """
    source_label = _DOC_SOURCE_LABELS.get(doc_source, doc_source)

    print("\n" + "!" * 70)
    print("PAUSED — newly extracted patient document needs your review")
    print("!" * 70)
    print(f"  Patient national ID : {patient_id}")
    print(f"  MDT request #        : {pre_request_id}")
    print(f"  Source                : {source_label}")
    print(f"  Saved to              : {pdf_path}")
    print("\n  This file was NOT already in your local folder — it was just pulled")
    print("  automatically for this row. Open it now and confirm it actually")
    print("  contains what this patient's request needs BEFORE it gets merged")
    print("  into the medical report + signed MDT form and uploaded.")
    print("!" * 70)

    while True:
        choice = input_fn(
            "\nType 'c' to CONTINUE (merge + upload this row), "
            "or 's' to SKIP this row and move to the next one: "
        ).strip().lower()
        if choice in ("c", "continue"):
            return "continue"
        if choice in ("s", "skip"):
            return "skip"
        print("  Please type 'c' or 's'.")


# =====================================================================
# BATCH PRE-SCAN (runs ONCE, before any MDT is created for ANY row)
# =====================================================================
#
# Replaces the old behaviour of resolving + pausing on a patient's
# document PDF one row at a time, in the middle of that row's own MDT
# creation. Instead, per your latest instruction:
#
#   1. Every unique patient ID in the whole queue is swept up front.
#      Anyone already sitting in PATIENT_DOCS_ROOT is left alone (just
#      refreshed with the last RECENT_ARCHIVE_DAYS_BACK days of CMIS
#      archive pages, same as before).
#   2. Anyone missing gets extracted RIGHT AWAY in this same pass -
#      SMC website first (merged with the same recent-archive window),
#      then the full CMIS archive as the final fallback - and saved into
#      PATIENT_DOCS_UNDER_PROCESSED_DIR instead of the cleaned root.
#   3. If anything landed in UNDER_PROCESSED, your labeling server is
#      launched and the run pauses for you to clean/label it, then
#      re-scans the cleaned root for the labeled result.
#   4. ONE global pause then asks whether to include those
#      freshly-extracted/labeled patients in this run at all ('c') or
#      skip them entirely - no MDT even created for them ('s').
#
# Only after all of that does the normal per-row loop (MDT creation,
# sign, report, merge, upload) begin.

def prescan_and_prepare_patient_documents(
    session: "SMCSession",
    unique_patient_ids: List[str],
    *,
    locate_fn=None,
) -> Dict[str, Dict]:
    """
    Runs locate_patient_document_pdf() once per unique patient ID up
    front (not interleaved with MDT creation). locate_fn is injectable
    for testing; defaults to the real locate_patient_document_pdf, which
    already implements local -> SMC website -> CMIS full archive in that
    order and already saves fallback extractions into
    FALLBACK_PATIENT_DOCS_DIR (== PATIENT_DOCS_UNDER_PROCESSED_DIR).

    Returns { patient_id: {"path": str|None, "source": str|None,
                            "newly_extracted": bool} }, one entry per ID
    in unique_patient_ids (order doesn't matter for the dict itself).
    """
    locate_fn = locate_fn or locate_patient_document_pdf
    doc_map: Dict[str, Dict] = {}
    total = len(unique_patient_ids)
    for idx, pid in enumerate(unique_patient_ids, start=1):
        log.info(f"\n[Pre-scan {idx}/{total}] Locating patient document PDF for {pid} …")
        try:
            path, source, newly = locate_fn(session, pid)
        except Exception as exc:
            log.warning(f"  Pre-scan lookup raised an error for {pid}, treating as not-found: {exc}")
            path, source, newly = None, None, False
        doc_map[pid] = {"path": path, "source": source, "newly_extracted": newly}
        if path:
            log.info(f"  -> [{source}] {path}")
        else:
            log.warning(f"  -> No patient document PDF found anywhere for {pid} "
                        "(local / SMC website / CMIS archive all came up empty).")
    return doc_map


def launch_doc_labeler_and_rescan(
    doc_map: Dict[str, Dict],
    *,
    find_local_fn=None,
    input_fn=input,
    launch_fn=None,
) -> Dict[str, Dict]:
    """
    Called once, right after prescan_and_prepare_patient_documents(),
    ONLY doing anything if at least one patient's document had to be
    freshly extracted (newly_extracted=True) - i.e. there's something
    sitting in PATIENT_DOCS_UNDER_PROCESSED_DIR that hasn't been
    cleaned/labeled yet. Rows whose document was already local
    (source == "local") are never touched here.

    Launches DOC_LABELER_BATCH_PATH (your labeling server's start3.bat,
    which already points --input at PATIENT_DOCS_UNDER_PROCESSED_DIR and
    --output at PATIENT_DOCS_ROOT), waits for you to confirm you're done,
    then re-searches PATIENT_DOCS_ROOT (the cleaned root, NOT
    UNDER_PROCESSED) for each freshly-extracted patient's ID so the rest
    of the pipeline picks up the CLEANED file instead of the raw one.

    launch_fn/find_local_fn/input_fn are injectable for offline testing.
    Mutates and returns doc_map.
    """
    find_local_fn = find_local_fn or find_patient_id_pdf
    launch_fn = launch_fn or _launch_doc_labeler_process

    newly = {pid: info for pid, info in doc_map.items() if info["newly_extracted"] and info["path"]}
    if not newly:
        log.info("\nNo freshly-extracted patient documents this run - nothing to label, "
                  "skipping the labeling-server step entirely.")
        return doc_map

    print("\n" + "=" * 70)
    print(f"{len(newly)} patient document PDF(s) were freshly extracted (SMC website / "
          "CMIS archive) this run and saved into:")
    print(f"  {PATIENT_DOCS_UNDER_PROCESSED_DIR}")
    for pid, info in newly.items():
        print(f"    {pid}  [{info['source']}]  ->  {info['path']}")
    print("=" * 70)

    if LAUNCH_DOC_LABELER_AUTOMATICALLY:
        launched = launch_fn(DOC_LABELER_BATCH_PATH)
        if launched:
            log.info(f"Launched the document labeling server: {DOC_LABELER_BATCH_PATH}")
        else:
            print(f"Could not launch it automatically - start it yourself: {DOC_LABELER_BATCH_PATH}")
    else:
        print(f"Start your labeling server yourself now: {DOC_LABELER_BATCH_PATH}")

    print(f"\nGo label/clean the file(s) above in the labeler. Each cleaned file must end up")
    print(f"saved under {PATIENT_DOCS_ROOT} itself (not the UNDER_PROCESSED subfolder),")
    print("named after the same patient ID - exactly what the labeler's own --output already does.")
    input_fn("\nPress Enter once you're done labeling/cleaning ALL of the file(s) listed above … ")

    log.info("Re-scanning the cleaned documents folder for the labeled file(s) …")
    for pid, info in newly.items():
        cleaned_path = find_local_fn(pid)
        if cleaned_path:
            log.info(f"  {pid}: found cleaned file -> {cleaned_path}")
            doc_map[pid]["path"] = cleaned_path
            doc_map[pid]["source"] = info["source"] + "+labeled"
        else:
            log.warning(f"  {pid}: still not found under {PATIENT_DOCS_ROOT} after labeling - "
                        f"keeping the raw UNDER_PROCESSED file ({info['path']}) as a fallback. "
                        "Check why the labeler didn't save a cleaned copy for this patient "
                        "before trusting this row.")
    return doc_map


def _launch_doc_labeler_process(batch_path: str) -> bool:
    """Actually starts the labeling server's .bat as a detached process so
    this script doesn't block on it. Returns False (never raises) if the
    path doesn't exist or the launch fails for any reason - the caller
    just tells you to start it yourself in that case."""
    if not batch_path or not os.path.exists(batch_path):
        log.warning(f"Labeling server batch file not found: {batch_path}")
        return False
    try:
        subprocess.Popen(["cmd.exe", "/c", "start", "", batch_path],
                          cwd=os.path.dirname(batch_path), shell=False)
        return True
    except Exception as exc:
        log.warning(f"Failed to launch labeling server ({batch_path}): {exc}")
        return False


def prompt_global_review_decision(doc_map: Dict[str, Dict], input_fn=input) -> str:
    """
    THE single, once-per-run pause (replaces the old per-row pause):
    decides whether rows whose patient document had to be freshly
    extracted this run (source != "local") are processed at all.

        'c' = continue  -> process every row, including the
                            freshly-extracted/newly-labeled ones.
        's' = skip      -> those rows are skipped ENTIRELY for this run
                            (no MDT is even created for them); only rows
                            whose document was already local from the
                            start are processed.

    Returns "continue" immediately, with NO prompt, if nothing was
    newly extracted this run (nothing to decide).
    """
    newly_ids = [pid for pid, info in doc_map.items() if info["newly_extracted"]]
    if not newly_ids:
        return "continue"

    print("\n" + "!" * 70)
    print("PAUSED - freshly extracted/labeled patient documents need your decision")
    print("!" * 70)
    print(f"  {len(newly_ids)} patient ID(s) had NO local document at the start of this run")
    print("  and were freshly pulled (SMC website / CMIS archive) and labeled:")
    for pid in newly_ids:
        info = doc_map[pid]
        print(f"    {pid}  [{info['source']}]  ->  {info['path']}")
    print("\n  Type 'c' to CONTINUE: process ALL of the rows above, in addition to every row")
    print("  whose document was already local from the start.")
    print("  Type 's' to SKIP: none of the rows above will be processed at all this run (no")
    print("  MDT is even created for them) - only rows with an already-local document proceed.")
    print("!" * 70)

    while True:
        choice = input_fn(
            "\nType 'c' to CONTINUE with all rows, or 's' to SKIP the freshly-extracted ones: "
        ).strip().lower()
        if choice in ("c", "continue"):
            return "continue"
        if choice in ("s", "skip"):
            return "skip"
        print("  Please type 'c' or 's'.")


def process_row(session: SMCSession, patient_id: str, description: str, tumor_type_raw: str,
                 request_category_raw: str, row_num: int, checkpoints: Dict[str, Dict],
                 precomputed_doc: Optional[Dict] = None,
                 medical_report_description: Optional[str] = None) -> Dict:
    result = {
        "row": row_num, "patient_id": patient_id, "description": description,
        "tumor_type": tumor_type_raw or "(blank -> Breast Cancer)",
        "status": "FAILED", "pre_request_id": "", "final_request_no": "",
        "output_path": "", "warnings": [], "error": "",
    }

    canonical, tumor_cfg_base = resolve_tumor_type(tumor_type_raw)
    if canonical is None:
        result["error"] = (
            f"Unrecognized tumor type {tumor_type_raw!r} — not found in TUMOR_TYPE_ALIASES. "
            "Add a mapping for it in the script before this row can be processed safely "
            "(diagnosis codes are not guessed at)."
        )
        return result

    request_category = resolve_request_category(request_category_raw)
    if request_category is None:
        result["error"] = (
            f"Unrecognized request category {request_category_raw!r} in Column D — expected "
            "blank/ordinary, 'surgery', or 'scan'. Add a mapping in REQUEST_CATEGORY_ALIASES "
            "before this row can be processed safely."
        )
        return result

    # tumor_cfg used for THIS row only: same diag_code/speciality_code as
    # the base tumor type, but proc_id may be overridden by surgery/scan.
    tumor_cfg = dict(tumor_cfg_base)
    tumor_cfg["proc_id"] = resolve_effective_proc_id(tumor_cfg_base, request_category)
    result["tumor_type"] = tumor_cfg["label"] + (
        f" [{request_category}]" if request_category != "ordinary" else ""
    )

    ckpt_key = f"{row_num}_{patient_id}"
    ckpt = checkpoints.get(ckpt_key, {})

    if ckpt.get("stage") == "uploaded":
        result["status"] = "SUCCESS"
        result["pre_request_id"] = ckpt.get("pre_request_id", "")
        result["final_request_no"] = ckpt.get("final_request_no", "")
        result["output_path"] = ckpt.get("merged_pdf_path", "")
        result["warnings"].append("Row already completed in a previous run — skipped re-processing.")
        return result

    if ckpt.get("stage") == "skipped_by_user":
        result["status"] = "SKIPPED_USER_REVIEW"
        result["pre_request_id"] = ckpt.get("pre_request_id", "")
        result["output_path"] = ckpt.get("reviewed_pdf_path", "")
        result["warnings"].append(
            "You previously chose to SKIP this row after reviewing its newly-extracted "
            f"document PDF ({ckpt.get('reviewed_pdf_path', '')}, via {ckpt.get('doc_source', '?')}). "
            "Delete this row's checkpoint entry if you want it re-attempted."
        )
        return result

    try:
        # ---- Stage 1: MDT creation ----
        if ckpt.get("stage") in ("mdt_created", "merged"):
            pre_request_id = ckpt["pre_request_id"]
            full_name = ckpt.get("full_name", "")
            log.info(f"  [resume] MDT already created previously: pre_request_id={pre_request_id}")
        else:
            log.info("  [Stage 1/6] Creating MDT request …")
            mdt_out = call_with_reconnect(session, "MDT creation", stage_create_mdt,
                                           session, patient_id, description, tumor_cfg)
            pre_request_id = mdt_out["pre_request_id"]
            full_name = mdt_out["full_name"]
            result["warnings"].extend(mdt_out["warnings"])
            log.info(f"    Created MDT request #{pre_request_id} for {patient_id}")
            save_checkpoint(CHECKPOINT_PATH, checkpoints, ckpt_key, {
                "stage": "mdt_created", "pre_request_id": pre_request_id,
                "full_name": full_name, "patient_id": patient_id,
            })
        result["pre_request_id"] = pre_request_id

        # ---- Stages 2-5: sign, report, find patient doc, merge ----
        if ckpt.get("stage") == "merged":
            merged_pdf_path = ckpt["merged_pdf_path"]
            log.info(f"  [resume] Merged PDF already built previously: {merged_pdf_path}")
        else:
            log.info("  [Stage 2/6] Rendering + signing MDT form …")
            mdt_signed_bytes = call_with_reconnect(session, "MDT render/sign", stage_render_and_sign,
                                                     session, pre_request_id, broad=True)

            log.info("  [Stage 3/6] Building medical report …")
            report_pdf_bytes = build_medical_report_pdf(
                full_name, patient_id,
                medical_report_description or description,
                tumor_cfg)

            log.info("  [Stage 4/6] Locating / refreshing patient document PDF …")
            if precomputed_doc is not None:
                # Batch mode (the normal path from main()): document lookup
                # + any labeling-server review already happened UP FRONT in
                # the pre-scan, for every row in the queue at once, before
                # any MDT was created. Nothing to look up or pause on here -
                # a row that shouldn't be processed at all (freshly
                # extracted + you chose 's') never reaches process_row in
                # the first place; see main().
                id_pdf_path = precomputed_doc.get("path")
                doc_source = precomputed_doc.get("source")
                newly_extracted = precomputed_doc.get("newly_extracted", False)
            else:
                # Standalone / backward-compatible path (e.g. calling
                # process_row() directly outside the normal main() batch
                # flow): resolve + pause right here, exactly as before.
                id_pdf_path, doc_source, newly_extracted = locate_patient_document_pdf(session, patient_id)

            if not id_pdf_path:
                result["error"] = (
                    f"No patient document PDF found under {PATIENT_DOCS_ROOT}, on the SMC "
                    f"website, or in the CMIS archive for patient {patient_id}. MDT request "
                    f"#{pre_request_id} was already created on the site — this row is skipped; "
                    "move on to the next one."
                )
                return result
            else:
                log.info(f"  Using patient document PDF ({doc_source}): {id_pdf_path}")

            # ---- Human review pause: only in the standalone path. In
            # batch mode this was already decided once, globally, for the
            # whole run before this row's MDT was even created - see
            # prompt_global_review_decision() / main(). ----
            if newly_extracted and precomputed_doc is None:
                decision = prompt_review_extracted_pdf(patient_id, id_pdf_path, doc_source,
                                                        pre_request_id)
                if decision == "skip":
                    result["status"] = "SKIPPED_USER_REVIEW"
                    result["output_path"] = id_pdf_path
                    result["warnings"].append(
                        f"Patient document PDF was newly extracted via {doc_source} and you "
                        f"chose to skip this row after reviewing it ({id_pdf_path}). MDT "
                        f"request #{pre_request_id} was already created on the site but this "
                        "row was NOT merged or uploaded."
                    )
                    save_checkpoint(CHECKPOINT_PATH, checkpoints, ckpt_key, {
                        "stage": "skipped_by_user", "pre_request_id": pre_request_id,
                        "full_name": full_name, "patient_id": patient_id,
                        "reviewed_pdf_path": id_pdf_path, "doc_source": doc_source,
                    })
                    return result
                log.info("  You chose to continue — proceeding to merge + upload this row.")

            log.info("  [Stage 5/6] Merging MDT + report + patient document …")
            merged_pdf_path = os.path.join(SUBMISSION_ROOT, f"{patient_id}_{pre_request_id}.pdf")
            merge_final_pdf(mdt_signed_bytes, report_pdf_bytes, id_pdf_path, merged_pdf_path)

            if not os.path.exists(merged_pdf_path) or os.path.getsize(merged_pdf_path) < 20_000:
                result["error"] = f"Merged PDF write failed or suspiciously small: {merged_pdf_path}"
                return result

            save_checkpoint(CHECKPOINT_PATH, checkpoints, ckpt_key, {
                "stage": "merged", "pre_request_id": pre_request_id,
                "full_name": full_name, "patient_id": patient_id,
                "merged_pdf_path": merged_pdf_path,
            })

        result["output_path"] = merged_pdf_path

        # ---- Stage 6: Upload ----
        log.info("  [Stage 6/6] Uploading merged PDF to the website …")
        final_request_no = call_with_reconnect(session, "Final upload", stage_upload_merged_pdf,
                                                 session, patient_id, pre_request_id, merged_pdf_path,
                                                 tumor_cfg)
        result["final_request_no"] = final_request_no
        result["status"] = "SUCCESS"
        log.info(f"    ✔ Success — final decree request number: {final_request_no}")

        save_checkpoint(CHECKPOINT_PATH, checkpoints, ckpt_key, {
            "stage": "uploaded", "pre_request_id": pre_request_id,
            "full_name": full_name, "patient_id": patient_id,
            "merged_pdf_path": merged_pdf_path, "final_request_no": final_request_no,
        })
        return result

    except RowBlacklisted as exc:
        result["status"] = "SKIPPED_BLACKLIST"
        result["warnings"].append(str(exc))
        return result
    except RowProcessingError as exc:
        result["error"] = str(exc)
        return result
    except RuntimeError as exc:
        # Exhausted reconnect retries.
        result["error"] = f"Network/reconnect failure: {exc}"
        return result
    except Exception as exc:
        log.exception(f"Row {row_num} ({patient_id}) failed unexpectedly")
        result["error"] = str(exc)
        return result


# =====================================================================
# EXCEL I/O
# =====================================================================

def load_queue(path: str) -> List[Tuple[str, str, str, str]]:
    """Returns (patient_id, description, tumor_type_raw, request_category_raw)
    per row. tumor_type_raw is '' when Column C is absent/blank -> Breast
    Cancer. request_category_raw is '' when Column D is absent/blank ->
    ordinary request (no surgery/scan proc_id override)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None:
            continue
        patient_id = str(row[0]).strip()
        description = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        tumor_type_raw = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        request_category_raw = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        if patient_id and description:
            rows.append((patient_id, description, tumor_type_raw, request_category_raw))
    return rows


def write_run_log(results: List[Dict], path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Run_Log"
    headers = ["Row", "Patient ID", "Tumor Type", "Description", "Status",
               "MDT PreRequest ID", "Final Request No", "Output Path", "Warnings", "Error"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center")

    status_colors = {
        "SUCCESS": "C6EFCE",
        "FAILED": "FFC7CE",
        "SKIPPED_BLACKLIST": "FFEB9C",
        "SKIPPED_USER_REVIEW": "FCE4D6",
    }
    for r in results:
        ws.append([
            r["row"], r["patient_id"], r["tumor_type"], r["description"], r["status"],
            r["pre_request_id"], r["final_request_no"], r["output_path"],
            "; ".join(r["warnings"]), r["error"],
        ])
        fill_color = status_colors.get(r["status"])
        if fill_color:
            ws.cell(row=ws.max_row, column=5).fill = PatternFill("solid", fgColor=fill_color)

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"
    wb.save(path)
    log.info(f"Run log written to: {path}")


# =====================================================================
# MAIN
# =====================================================================

def main():
    if not os.path.exists(EXCEL_INPUT_PATH):
        sys.exit(f"Input Excel not found: {EXCEL_INPUT_PATH}")
    from medical_report_overlay import MEDICAL_REPORT_TEMPLATE_PDF
    if not MEDICAL_REPORT_TEMPLATE_PDF.exists():
        sys.exit(
            f"Medical report base PDF not found: {MEDICAL_REPORT_TEMPLATE_PDF}\n"
            "Export medical_report_template.docx to PDF once (Word -> Save As PDF) "
            "and point MEDICAL_REPORT_TEMPLATE_PDF (in medical_report_overlay.py) at it."
        )
    if not os.path.exists(WKHTMLTOPDF_PATH):
        sys.exit(f"wkhtmltopdf not found at: {WKHTMLTOPDF_PATH} — install it and update WKHTMLTOPDF_PATH.")
    if not os.path.isdir(PATIENT_DOCS_ROOT):
        sys.exit(f"Patient documents folder not found: {PATIENT_DOCS_ROOT}")

    queue = load_queue(EXCEL_INPUT_PATH)
    if not queue:
        sys.exit("No valid rows (ID + description) found in the input Excel.")
    log.info(f"Loaded {len(queue)} row(s) from queue.")

    checkpoints = load_checkpoints(CHECKPOINT_PATH)
    if checkpoints:
        done = sum(1 for v in checkpoints.values() if v.get("stage") == "uploaded")
        log.info(f"Loaded {len(checkpoints)} checkpoint(s) from a previous run ({done} already fully uploaded).")

    session = SMCSession()
    if not session.login():
        sys.exit("Login failed. Aborting.")

    # ---- BATCH PRE-SCAN (new, once per run, before any MDT creation) ----
    # 1. Sweep every unique patient ID in the queue: local folder first
    #    (refreshed with the last RECENT_ARCHIVE_DAYS_BACK days of CMIS
    #    archive pages), then SMC website, then full CMIS archive - for
    #    IDs with no local file. Fallback extractions land in
    #    PATIENT_DOCS_UNDER_PROCESSED_DIR, not the cleaned root.
    unique_patient_ids = list(dict.fromkeys(pid for pid, *_ in queue))
    log.info(f"\n{'='*65}\nPre-scan: locating patient document PDFs for all "
             f"{len(unique_patient_ids)} unique patient ID(s) in the queue before "
             f"creating any MDT requests …\n{'='*65}")
    doc_map = prescan_and_prepare_patient_documents(session, unique_patient_ids)

    # 2. If anything had to be freshly extracted, hand it to your labeling
    #    server and wait for you to clean/label it, then re-find the
    #    cleaned copies under PATIENT_DOCS_ROOT.
    doc_map = launch_doc_labeler_and_rescan(doc_map)

    # 3. ONE global pause: include the freshly-extracted/labeled patients
    #    in this run ('c'), or skip them entirely, no MDT created ('s').
    #    No prompt at all if every patient's document was already local.
    global_decision = prompt_global_review_decision(doc_map)
    skip_patient_ids = set()
    if global_decision == "skip":
        skip_patient_ids = {pid for pid, info in doc_map.items() if info["newly_extracted"]}
        log.info(f"\nSkipping {len(skip_patient_ids)} freshly-extracted patient ID(s) entirely "
                 f"this run (no MDT will be created for them): {sorted(skip_patient_ids)}")

    results: List[Dict] = []

    for i, (patient_id, description, tumor_type_raw, request_category_raw) in enumerate(queue, start=1):
        log.info(f"\n{'='*65}\nProcessing row {i}/{len(queue)}: {patient_id}  "
                 f"(tumor type: {tumor_type_raw or 'Breast Cancer (default)'}"
                 f"{', category: ' + request_category_raw if request_category_raw else ''})\n{'='*65}")

        if patient_id in skip_patient_ids:
            doc_info = doc_map.get(patient_id, {})
            log.info(f"  Skipping entirely — document was freshly extracted this run and you "
                     f"chose 's' at the global review pause. No MDT created for this row.")
            result = {
                "row": i, "patient_id": patient_id, "description": description,
                "tumor_type": tumor_type_raw or "(blank -> Breast Cancer)",
                "status": "SKIPPED_USER_REVIEW", "pre_request_id": "", "final_request_no": "",
                "output_path": doc_info.get("path", ""), "warnings": [
                    f"Patient document PDF was freshly extracted this run via "
                    f"{doc_info.get('source', '?')} ({doc_info.get('path', '?')}) and you chose "
                    "'s' (skip) at the once-per-run global review pause — no MDT was created "
                    "for this row at all."
                ], "error": "",
            }
            results.append(result)
            write_run_log(results, RUN_LOG_PATH)
            continue

        result = process_row(session, patient_id, description, tumor_type_raw,
                              request_category_raw, i, checkpoints,
                              precomputed_doc=doc_map.get(patient_id))
        results.append(result)

        log.info(
            f"  Status: {result['status']}"
            + (f" | MDT #{result['pre_request_id']}" if result["pre_request_id"] else "")
            + (f" | Final request #{result['final_request_no']}" if result["final_request_no"] else "")
            + (f" | ERROR: {result['error']}" if result["error"] else "")
        )
        for w in result["warnings"]:
            log.info(f"  [warning] {w}")

        write_run_log(results, RUN_LOG_PATH)

        if i == 1 and PAUSE_AFTER_FIRST_ROW:
            print("\n" + "=" * 65)
            print("First row processed. Check the merged PDF, the MDT form on the")
            print("site, and the run log before continuing:")
            print(f"  Output PDF : {result.get('output_path') or '(none — see error)'}")
            print(f"  Run log    : {RUN_LOG_PATH}")
            print("=" * 65)
            input("Press Enter to continue with the remaining rows, or Ctrl+C to stop … ")

    succeeded = sum(1 for r in results if r["status"] == "SUCCESS")
    skipped_bl = sum(1 for r in results if r["status"] == "SKIPPED_BLACKLIST")
    skipped_review = sum(1 for r in results if r["status"] == "SKIPPED_USER_REVIEW")
    failed = sum(1 for r in results if r["status"] == "FAILED")
    log.info(
        f"\nDone. {succeeded}/{len(results)} row(s) fully submitted. "
        f"{skipped_bl} blacklisted, {skipped_review} skipped after your review of a newly "
        f"extracted document, {failed} failed.\n"
        f"Run log: {RUN_LOG_PATH}\nCheckpoints: {CHECKPOINT_PATH}"
    )


# =====================================================================
# SELF-TEST — offline verification of the Stage-4 fallback order +
# review-pause logic, with NO network calls, NO real credentials, and
# NO real files. Run this first: python Unified_Decree_Submission_Pipeline.py --self-test
# =====================================================================

def run_self_test():
    """
    Exercises locate_patient_document_pdf() and
    prompt_review_extracted_pdf() against fake stand-ins for every
    network/filesystem dependency, and asserts the exact control-flow
    you asked for. Prints PASS/FAIL per case and a final summary; exits
    non-zero if anything fails, so it's safe to script/CI this before a
    real run.
    """
    print("=" * 70)
    print("SELF-TEST: Stage 4 fallback order + review-pause logic (offline)")
    print("=" * 70)

    failures = []

    def check(label: str, condition: bool, detail: str = ""):
        status = "PASS" if condition else "FAIL"
        print(f"  [{status}] {label}" + (f" — {detail}" if detail and not condition else ""))
        if not condition:
            failures.append(label)

    fake_session = object()  # locate_patient_document_pdf never touches it directly;
                              # only the injected reconnect_fn sees it.

    def fake_reconnect(session, stage_label, func, *args, **kwargs):
        # Stand-in for call_with_reconnect(): just call straight through,
        # no real retry/reconnect machinery needed for this offline test.
        return func(*args, **kwargs)

    # ---- Case 1: found locally -> used as-is, no pause, source="local" ----
    def local_found(_pid):
        return "/fake/local/12345.pdf"

    def refresh_says_nothing_new(_pid, _path, **_kwargs):
        return False

    path, source, newly = locate_patient_document_pdf(
        fake_session, "12345",
        find_local_fn=local_found, refresh_fn=refresh_says_nothing_new,
        website_fn=lambda *a, **k: (_ for _ in ()).throw(AssertionError("website_fn should not be called")),
        full_archive_fn=lambda *a, **k: (_ for _ in ()).throw(AssertionError("full_archive_fn should not be called")),
        reconnect_fn=fake_reconnect,
    )
    check("Case 1: local file found -> used directly", path == "/fake/local/12345.pdf")
    check("Case 1: source reported as 'local'", source == "local")
    check("Case 1: NOT flagged newly_extracted (no pause)", newly is False)

    # ---- Case 2: not local, found on website, CMIS has a same-day merge ----
    def local_missing(_pid):
        return None

    def website_finds_it(*a, **k):
        return "/fake/website/12345.pdf"

    merge_calls = []

    def refresh_merges_something(pid, path_, **_kwargs):
        merge_calls.append((pid, path_))
        return True

    path, source, newly = locate_patient_document_pdf(
        fake_session, "12345",
        find_local_fn=local_missing, website_fn=website_finds_it,
        refresh_fn=refresh_merges_something,
        full_archive_fn=lambda *a, **k: (_ for _ in ()).throw(AssertionError("full_archive_fn should not be called")),
        reconnect_fn=fake_reconnect,
    )
    check("Case 2: SMC website tried BEFORE CMIS full archive", path == "/fake/website/12345.pdf")
    check("Case 2: source reported as 'website+archive'", source == "website+archive")
    check("Case 2: IS flagged newly_extracted (pause required)", newly is True)
    check("Case 2: CMIS recent-archive merge was attempted on the website PDF",
          merge_calls == [("12345", "/fake/website/12345.pdf")])

    # ---- Case 3: not local, found on website, nothing new in CMIS ----
    path, source, newly = locate_patient_document_pdf(
        fake_session, "12345",
        find_local_fn=local_missing, website_fn=website_finds_it,
        refresh_fn=refresh_says_nothing_new,
        full_archive_fn=lambda *a, **k: (_ for _ in ()).throw(AssertionError("full_archive_fn should not be called")),
        reconnect_fn=fake_reconnect,
    )
    check("Case 3: website file still used when CMIS has nothing new", path == "/fake/website/12345.pdf")
    check("Case 3: still flagged newly_extracted", newly is True)

    # ---- Case 4: not local, website empty -> full CMIS archive extraction ----
    def website_finds_nothing(*a, **k):
        return None

    def full_archive_finds_it(_pid, _dir):
        return "/fake/cmis_archive/12345_archive.pdf"

    path, source, newly = locate_patient_document_pdf(
        fake_session, "12345",
        find_local_fn=local_missing, website_fn=website_finds_nothing,
        refresh_fn=refresh_says_nothing_new, full_archive_fn=full_archive_finds_it,
        reconnect_fn=fake_reconnect,
    )
    check("Case 4: falls through to full CMIS archive when website is empty",
          path == "/fake/cmis_archive/12345_archive.pdf")
    check("Case 4: source reported as 'cmis_full_archive'", source == "cmis_full_archive")
    check("Case 4: IS flagged newly_extracted (pause required)", newly is True)

    # ---- Case 5: nothing anywhere ----
    def full_archive_finds_nothing(_pid, _dir):
        return None

    path, source, newly = locate_patient_document_pdf(
        fake_session, "12345",
        find_local_fn=local_missing, website_fn=website_finds_nothing,
        refresh_fn=refresh_says_nothing_new, full_archive_fn=full_archive_finds_nothing,
        reconnect_fn=fake_reconnect,
    )
    check("Case 5: nothing found anywhere -> path is None", path is None)
    check("Case 5: source is None", source is None)
    check("Case 5: not flagged newly_extracted (nothing to review)", newly is False)

    # ---- Case 6: review pause — typing 'c' continues ----
    fake_inputs_c = iter(["c"])
    decision = prompt_review_extracted_pdf(
        "12345", "/fake/website/12345.pdf", "website+archive", "PR-999",
        input_fn=lambda _prompt: next(fake_inputs_c),
    )
    check("Case 6: typing 'c' returns 'continue'", decision == "continue")

    # ---- Case 7: review pause — typing 's' skips ----
    fake_inputs_s = iter(["s"])
    decision = prompt_review_extracted_pdf(
        "12345", "/fake/cmis_archive/12345_archive.pdf", "cmis_full_archive", "PR-999",
        input_fn=lambda _prompt: next(fake_inputs_s),
    )
    check("Case 7: typing 's' returns 'skip'", decision == "skip")

    # ---- Case 8: review pause — re-prompts on garbage input, then accepts 'skip' ----
    fake_inputs_retry = iter(["blah", "", "skip"])
    decision = prompt_review_extracted_pdf(
        "12345", "/fake/website/12345.pdf", "website+archive", "PR-999",
        input_fn=lambda _prompt: next(fake_inputs_retry),
    )
    check("Case 8: invalid input is rejected and it re-prompts until valid", decision == "skip")

    # ---- Case 9: batch pre-scan calls locate_fn once per unique ID ----
    def fake_locate(_session, pid, **_kwargs):
        if pid == "111":
            return "/fake/local/111.pdf", "local", False
        if pid == "222":
            return "/fake/under_processed/222.pdf", "website+archive", True
        return None, None, False

    doc_map = prescan_and_prepare_patient_documents(
        fake_session, ["111", "222", "333"], locate_fn=fake_locate,
    )
    check("Case 9: pre-scan returns one entry per unique patient ID", len(doc_map) == 3)
    check("Case 9: local patient reported not newly_extracted", doc_map["111"]["newly_extracted"] is False)
    check("Case 9: freshly-extracted patient reported newly_extracted", doc_map["222"]["newly_extracted"] is True)
    check("Case 9: not-found patient has path=None", doc_map["333"]["path"] is None)

    # ---- Case 10: labeler step is skipped entirely when nothing was freshly extracted ----
    launch_calls = []
    doc_map_local_only = {"111": {"path": "/fake/local/111.pdf", "source": "local", "newly_extracted": False}}
    result_map = launch_doc_labeler_and_rescan(
        doc_map_local_only,
        find_local_fn=lambda _pid: (_ for _ in ()).throw(AssertionError("should not re-scan")),
        input_fn=lambda _p: (_ for _ in ()).throw(AssertionError("should not prompt")),
        launch_fn=lambda _path: launch_calls.append(_path) or True,
    )
    check("Case 10: labeler NOT launched when nothing newly extracted", launch_calls == [])
    check("Case 10: doc_map returned unchanged", result_map is doc_map_local_only)

    # ---- Case 11: labeler step launches, waits, and re-scans for freshly extracted docs ----
    doc_map_fresh = {"222": {"path": "/fake/under_processed/222.pdf", "source": "website+archive", "newly_extracted": True}}
    launch_calls2 = []
    waited = []
    result_map2 = launch_doc_labeler_and_rescan(
        doc_map_fresh,
        find_local_fn=lambda pid: "/fake/cleaned/222.pdf" if pid == "222" else None,
        input_fn=lambda _p: waited.append(True),
        launch_fn=lambda path: launch_calls2.append(path) or True,
    )
    check("Case 11: labeler launched exactly once", launch_calls2 == [DOC_LABELER_BATCH_PATH])
    check("Case 11: run paused waiting for confirmation", waited == [True])
    check("Case 11: cleaned path picked up after re-scan", result_map2["222"]["path"] == "/fake/cleaned/222.pdf")
    check("Case 11: source annotated with '+labeled'", result_map2["222"]["source"] == "website+archive+labeled")

    # ---- Case 12: global review decision — 'c' includes freshly-extracted rows ----
    decision = prompt_global_review_decision(doc_map_fresh, input_fn=lambda _p: "c")
    check("Case 12: 'c' returns 'continue'", decision == "continue")

    # ---- Case 13: global review decision — 's' means skip them ----
    decision = prompt_global_review_decision(doc_map_fresh, input_fn=lambda _p: "s")
    check("Case 13: 's' returns 'skip'", decision == "skip")

    # ---- Case 14: global review decision — no prompt at all when nothing was freshly extracted ----
    decision = prompt_global_review_decision(
        doc_map_local_only,
        input_fn=lambda _p: (_ for _ in ()).throw(AssertionError("should not prompt — nothing to decide")),
    )
    check("Case 14: no freshly-extracted docs -> auto 'continue', no prompt", decision == "continue")

    print("=" * 70)
    if failures:
        print(f"SELF-TEST FAILED — {len(failures)} check(s) did not pass:")
        for f in failures:
            print(f"    - {f}")
        print("Fix these before running against real patients/live sites.")
        sys.exit(1)
    else:
        print(f"SELF-TEST PASSED — all checks green. Safe to run for real.")
    print("=" * 70)


if __name__ == "__main__":
    if "--self-test" in sys.argv:
        run_self_test()
    else:
        main()
